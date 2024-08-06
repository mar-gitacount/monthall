# -*- coding: utf-8 -*-
from selenium import webdriver
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
import os
import sys
import logging


from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

from selenium.webdriver.chrome.options import Options

num = int(sys.argv[1]) if len(sys.argv) > 1 else 0

# ここで num を使用して何かを実行する

(f"受け取った引数: {num}")
url = f"https://www.909.co.jp/rolex_search/result.html?p={num}&sort=DESC&snum=16"
url = f"https://www.gmt-j.com/search?order=pop&rows=100&itemwear=0&installments=0&specpanelopen=0&maker=281&category=337,339,338&stockstatus=1,2,3#itemlis"
url = f"https://www.gmt-j.com/search?order=pop&rows=100&page={num}&itemwear=0&installments=0&specpanelopen=0&maker=281&category=337,339,338&stockstatus=1,2,3#itemlist"
# url = f"https://kakaku.com/watch_accessory/watch/itemlist.aspx?pdf_ma=5090&pdf_Spec106=1,2,4&pdf_vi=c&pdf_pg={num}"


# ?モデル、ブレスレット、文字盤を抽出する関数
def model_validete_imput(text):
    models = [
        "デイトジャスト",
        "オイスター",
        "コスモグラフ",
        "シードゥエラー",
        "エクスプローラー",
        "GMTマスター",
        "GMTマスターII",
        "サブマリーナー",
        "ヨットマスター",
        "スカイドゥエラー",
        "エクスプローラーII",
        "エアキング",
    ]
    print("関数内のテキスト", text)
    pattern = r"\b\s+(\S+)\s+\b"
    beltpattern = r"\[(.*?)\]|\((.*?)\)"
    # ベルトと文字盤を抽出する。
    beltmatches = re.findall(beltpattern, text)
    # モデル名を抽出する。
    model = re.sub(beltpattern, "", text)
    print("モデル名", model)
    # [],()を正規表現で抽出する。
    items = {"model": model, "beltmatches": beltmatches}
    return items


def find_all_matches(pattern, text):
    # 正規表現にマッチしたすべての部分文字列を配列として取得
    matches = re.findall(pattern, text)
    return matches


def join_tuples(tuple_list):
    # タプルの要素をすべて連結して一つの文字列にする
    return "".join(item for tpl in tuple_list for item in tpl)


# ?エクセルに入力する関数
def wsinsert(values, sheet):
    print("wsinsert関数", values)
    sheet.append(values)
    # for item in values:
    #     sheet.append(item)


# def tag_uncle_item_get(tag, text):
#     getitem = item_soup.find(tag, text=text)
#     uncle_item = ""
#     if getitem:
#         # 親タグ取得
#         parent_element = getitem.parent
#         if parent_element:
#             # 親兄弟タグ取得
#             next_sibling = parent_element.find_next_sibling().get_text(strip=True)
#             uncle_item = next_sibling
#             return uncle_item
#     return uncle_item


# 現在の日付を取得
today_date = datetime.now().strftime("%Y%m%d")
# ファイル名に日付を組み込む

file_name = f"GMT{today_date}.xlsx"
if not os.path.exists(file_name):
    # Excelブックの作成
    wb = Workbook()
    ws = wb.active
    # ヘッダー行を追加
    ws.append(
        [
            "モデル名",
            "リファレンスNO",
            "ダイヤル",
            "ブレスレット",
            "値段",
            "URL",
        ]
    )
else:
    # ファイルが存在する場合は既存のファイルを読み込み
    wb = load_workbook(file_name)
    ws = wb.active


options = Options()
options.add_argument("--headless")  # ヘッドレスモードを有効にする


# SeleniumのWebDriverを初期化
driver = webdriver.Chrome(options=options)  # または他のブラウザに合わせて選択

# URLを開く
driver.get(url)

# ページが完全に読み込まれるまで待つ
wait = WebDriverWait(driver, 10)
wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))


# Seleniumがページのロードを待つなどの適切な待機処理が必要な場合はここで実施

# ページのHTMLを取得
page_source = driver.page_source

# BeautifulSoupを使ってHTMLを解析
soup = BeautifulSoup(page_source, "html.parser")

# ログの設定
logging.basicConfig(filename=f"{today_date}error.log", level=logging.ERROR)


# !ここから処理スタート
# <tbody> タグ内のテキストを抽出して表示
tbody_tag = soup.find("body")
target_tag = tbody_tag.find("FL-wrap")
resultBox_items = tbody_tag.find("div", id="resultBox")
# リファレンスナンバーを取得する
ref_pattern = r"\b\d{3,}[A-Z]*\b"
# ref_pattern = r".*?\b\d{3,}[A-Z]*\b"


rolex_pattern = r"ロレックス(.+)"
rolex_alphabet_pattern = r"ROLEX(.+)"
rolex_pattern = r"ROLEX ロレックス"


# ブレスレットパターンを指定
# blesslet_pattern = r"\s*ブレスレット\s*(.+)|\s*(.+)\s*ブレスレット"
blesslet_pattern = r"\s+ブレスレット\s*(.+)"

# 色の正規表現
color_pattern = r"\b\s+(.+)*(レッド|ブルー|ブラック|イエロー|グリーン|ダークグレー|ホワイト|シャンパン|シルバー|ゴールド|アイスブルー|ブラウン|スレートローマ)+(.*)\b"
color_pattern = r"\b(/)*(ラピスラズリグリーンパーム|チョコレートコンピューター|オーベルジーヌ VI IXダイヤ|ゴールデンフルーテッドモチーフ|ソーダライト|ロゼ|スパイダーダイアル|オールトリチウム|トリチノバ|オリーブグリーン|パヴェダイヤ|シャンパンフルーテッド|ブラックシェル|NEWダイヤル|ロゼローマ |スチール|ブルーローマ|新ブルーダイヤル|ホワイトシェル|シルバー スターダイヤ |チョコレート|ムーンフェイズ|スレート|Dブルー|アイボリー|ブライトブルー|ターコイズブルーセレブレーション モチーフ|オイスター|キャンディピンク|ホワイトローマ|ダークロジウム|ターコイズブルーセレブレーション|ターコイズブルー|セレブレーションモチーフ|ブラック コンセントリック|コーラルレッド|レッド|ブルー|ブラック|イエロー|グリーン|グレー|ダークグレー|ホワイト|シャンパン|シルバー|ゴールド|アイスブルー|ブラウン|スレートローマ|ブルーブラック)\b"

# ?ブレスレット名の後に続く場合のパターン
blesslet_pattern = r".*ブレスレット\s*(.+)"

# blesslet_pattern = r"\s+(.+)ブレスレット"
blesslet_pattern = r"\s+(.+)\s+(.+)\s+(.+)ブレスレット"
blesslet_pattern = r"(.+?)\s(.+)ブレスレット"
# ブレスレットを抽出する
blesslet_pattern = r"\s+(.+?)ブレスレット"
blesslet_pattern2 = r"\s+(.+?)ブレス"
blesslet_pattern = r"\s+(.+?)ブレス(?:レット)*.*"

# 空白のパターンもちゅうしゅつ
blank_pattern = r"\s+(\S+)"
# price_get = resultBox_items.find_all("p",class_="FP-resultItem__price")

# print(resultBox_items)
items = resultBox_items.find_all("div", class_="FP-resultItem__textBox")
for item in items:

    a_tag_get = item.find("a")
    # aタグを取得する
    href = a_tag_get.get("href")
    # FP-resultItem__price

    href = "https://www.gmt-j.com" + href
    print(href)
    itemname = item.find("h2", class_="FP-resultItem__name").get_text(strip=True)
    print(itemname)
    try:
        ref = re.findall(ref_pattern, itemname)[0]
    except Exception as e:
        ref_pattern = r"\b(\d{3,}[A-Z]*)\b"
        # ref = re.findall(ref_pattern, itemname)
        match = re.search(ref_pattern, itemname)
        if match:
            print("マッチ")
        else:
            print("マッチしない")
            ref = "正しいリファレンスナンバーを入力してください"

    print(ref, "リファレンス")
    itemname = re.sub(rolex_pattern, "", itemname)
    itemname = re.sub(ref_pattern, "", itemname)
    blesslet = re.findall(blesslet_pattern, itemname)

    if blesslet:
        # blesslet = blesslet.group(1).strip()
        # pattern = r"\s+(.+)\b"
        print(blesslet, "があります！")
        blank_get = re.findall(blank_pattern, blesslet[0])
        match = re.search(r"\s+(\S+)\s*$", blesslet[0])
        if match:
            # ブレスレットを抽出する。
            blessletitem = match.group(1)
            # ブレスレットのアイテムを削除する。
            blesslet_remove = re.sub(blessletitem, "", blesslet[0])
            # ブレスレットを除いた値
            itemname = blesslet_remove
            blesslet = blessletitem
            print(blesslet_remove, "これはブレスレットを削除した値")
    else:
        print("ブレスレットがない")
        # blesslet = re.findall(blesslet_pattern2, itemname)
        blesslet = ""
        print(blesslet, "はブレスのパターンです")
    # カラーを抽出する。
    colorcheck = re.findall(color_pattern, itemname)

    # colorcheck = find_all_matches(color_pattern, itemname)
    if colorcheck:
        print(colorcheck, "色を抽出しました")
        # colorcheckをすべて連結させる
        # coloritem = "".join(colorcheck)
        coloritem = join_tuples(colorcheck)
        print(coloritem, "色です")
        itemname = re.sub(coloritem, "", itemname)
    else:
        print("色がない")
        coloritem = ""
    # !最後はアイテムネームのみが残る。
    print(itemname)
    try:
        price_get = item.find("p", class_="FP-resultItem__price").get_text(strip=True)
    except Exception as e:
        price_get = item.find("p", class_="FP-resultItem__priceHighlight").get_text(
            strip=True
        )
        print(price_get)

    print(ref)
    ws.append([itemname, ref, coloritem, blesslet, price_get, href])
    # 各ページへアクセスする。
    # driver.get(href)
    # item_page_source = driver.page_source
    # item_soup = BeautifulSoup(item_page_source, "html.parser")
    # h1 = item_soup.find("h1", class_="clearfix").getText(strip=True)
    # print(h1)
wb.save(file_name)


# tbody_tag = soup.find("tbody")


# table_get = tbody_tag.find("div", class_="list_container")
# a_tag_get = table_get.find_all("a", class_="list_item")
# try:
#     a_tag_get = table_get.find_all("a", class_="list_item")
# except Exception as e:
#     # エラーログを記録
#     logging.error(f"An error occurred: {e}")

# try:
#     for item in a_tag_get:

#         href = item.get("href")

#         href = href.replace("..", "")
#         href = "https://www.909.co.jp" + href
#         print(href)
#         driver.get(href)
#         item_page_source = driver.page_source
#         item_soup = BeautifulSoup(item_page_source, "html.parser")
#         # ?アイテム名とリファレンスナンバー
#         # 新品
#         itemnameandrefno_element = item_soup.find("h2", class_="new")
#         # 中古
#         itemnameandrefno_element_used = item_soup.find("h2", class_="used")
#         #
#         itemnameandrefno_element_salonused = item_soup.find("h2", class_="salon_used")
#         # ビンテージ
#         itemnameandrefno_element_vin = item_soup.find("h2", class_="vin")
#         # プレミアムモダン
#         itemnameandrefno_element_pm = item_soup.find("h2", class_="pm")

#         if itemnameandrefno_element:
#             itemnameandrefno = itemnameandrefno_element.get_text(strip=True)
#         elif itemnameandrefno_element_used:
#             itemnameandrefno = itemnameandrefno_element_used.get_text(strip=True)
#         elif itemnameandrefno_element_vin:
#             itemnameandrefno = itemnameandrefno_element_vin.get_text(strip=True)
#         elif itemnameandrefno_element_pm:
#             itemnameandrefno = itemnameandrefno_element_pm.get_text(strip=True)
#         elif itemnameandrefno_element_salonused:
#             itemnameandrefno = itemnameandrefno_element_salonused.get_text(strip=True)
#         else:
#             itemnameandrefno_element = ""
#         # リファレンスナンバーを取得する
#         ref_pattern = r"Ref\.\d+[A-Z]*"

#         # アイテム名を取得する
#         item_pattern = r"ロレックス(.+)"

#         ref_no = re.findall(ref_pattern, itemnameandrefno)[0]
#         ref_no = ref_no.replace("Ref.", "")
#         # itemname = re.sub(ref_pattern, "", itemnameandrefno)
#         itemname = re.findall(item_pattern, itemnameandrefno)[0]
#         itemname = re.sub(ref_pattern, "", itemname)
#         # ?金額
#         price = item_soup.find("span", class_="sell_price").get_text(strip=True)
#         # ?ダイヤル
#         dial = item_soup.find("b", string="ダイヤル")
#         if dial:
#             # 親タグ
#             parent_element = dial.parent
#             if parent_element:
#                 next_sibling = parent_element.find_next_sibling().get_text(strip=True)
#                 dial_text = next_sibling
#             else:
#                 dial_text = ""

#         # ?ブレスレット
#         blesslet = tag_uncle_item_get("b", "ブレスタイプ")
#         # ?ケース径
#         case_size = tag_uncle_item_get("b", "ケース径")
#         ws.append([itemname, ref_no, dial_text, blesslet, price, href])
#         print(blesslet, "はブレスレット")
#         print(itemnameandrefno, "アイテムとリファレンスナンバー")
#         print(price, "金額")
#         print(dial_text, "ダイヤル")
#         print(blesslet, "ブレスレット")
#         print(case_size, "ケースサイズ")
# except Exception as e:
#     # エラーログを記録
#     logging.error(f"An error occurred: {e}its url {url}")

# # ここでエクセル保存する。
# wb.save(file_name)
