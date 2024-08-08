from openpyxl import Workbook, load_workbook
import os
import re
from datetime import datetime
import shutil
from whochedata_sqlite_data_insert import WhocheSqliteDataInsert


def copy_folder(src_folder, dest_folder):
    if not os.path.exists(dest_folder):
        os.makedirs(dest_folder)
    
    try:
        shutil.copytree(src_folder, dest_folder, dirs_exist_ok=True)
        print(f"Folder copied from {src_folder} to {dest_folder}")
    except Exception as e:
        print(f"Error: {e}")

# ウォッチニアン(買取)のデータをデータベースに入稿する処理
def watchnianbuy(instance):
    current_directory = os.getcwd()
    folder_name = "まとめるエクセル"
    new_directory = os.path.join(current_directory, folder_name)
    items = instance.tablesdateill
    watch_item_instance = items["watch_item"]
    # pattern = re.compile(r'^ウォッチニアン.*\.xlsx$')
    pattern = re.compile(r'^ウォッチニアン.*\買取.*\.xlsx$')
    # 買取正規表現
    buygrammerpattern = r"(買取?)"
    salegrammerpattern = r"(販売)"
    numpattern = r"\d+"
    datepattern = r'(\d{4}).*(\d{2})'
    wachnian_files =[]
    for file in os.listdir(new_directory):
        if pattern.match(file):
            wachnian_files.append(os.path.join(new_directory, file))
    for file in wachnian_files:
        full_path = os.path.join(new_directory, file)
        workbook = load_workbook(full_path, data_only=False)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                rowitemslist = list(row)
                model_name = rowitemslist[0]
                ref = rowitemslist[1]
                year = ""
                size =""
                dial = rowitemslist[2]
                bracelet = ""
                new_price = rowitemslist[3]
                used_price = rowitemslist[4]
                unused_price = rowitemslist[5]

                print(new_price)
                new_price = re.findall(numpattern,str(new_price))
                used_price = re.findall(numpattern,str(used_price))
                unused_price = re.findall(numpattern,str(unused_price))
                 
                try:
                    if isinstance(new_price[0],tuple):
                        new_price = "".join(new_price[0])
                except IndexError as e:
                    print("文字列or配列")
                except IndexError as e:
                    print("test")
                try:
                    if isinstance(used_price[0],tuple):
                        used_price = "".join(used_price[0])
                except IndexError as e:
                    print("文字列or配列")
                except IndexError as e:
                    print("test")
                try:
                    if isinstance(unused_price[0],tuple):
                        unused_price = "".join(unused_price[0])
                except IndexError as e:
                    print("文字列or配列")
                except IndexError as e:
                    print("test")
                
                if new_price:
                    print(f"{new_price}はエラーになる値段")
                    new_price = "".join(new_price)
                    new_price = int(new_price)
                else:
                    new_price = 0
                
                if used_price:
                    used_price = "".join(used_price)
                    used_price = int(used_price)
                else:
                    used_price = 0

                if unused_price:
                    unused_price = "".join(unused_price)
                    unused_price = int(unused_price)
                else:
                    unused_price = 0
                url = rowitemslist[6]

                if not url:
                    url = rowitemslist[6]
                match = re.search(r'https://buy.watchnian.com/.+_(\d+)', url)
                print(url)
                if match:
                    product_id = match.group(1)
                else:
                    print("----------------")
                    continue
               
                
         
                print(f"リファレンス:{ref} 年:{year} サイズ{size} ダイアル{dial} ブレスレット{bracelet} URL{url} 新品値段{new_price} 中古値段{used_price}")
                yearmonthpattern = re.compile(r'_(\d{4}-\d{2})')
                # ファイル名
                filename = os.path.basename(file)
                datematches = re.findall(datepattern,filename)
                print(datematches)
                dateresult = "/".join(datematches[0])
                make_ID_Format_Date = dateresult.replace("/","-")

                make_ID = "Watchnian-"+make_ID_Format_Date+"-"+product_id
                # 買取、販売チェックする。

                yearmonthmatch = re.search(r'_*(\d{4}-\d{2})', filename)
                # yearmonthmatch = yearmonthpattern.match(str(filename))
                buy_match = re.search(buygrammerpattern,filename)
                # 
                if buy_match:
                    otherdata = buy_match.group(1)
                else:
                    otherdata = "販売"
               
               
               
                if yearmonthmatch:
                    formatted_date = yearmonthmatch.group(1)  # 年月を取得する
                    formatted_date = datetime.strptime(formatted_date, '%Y-%m')
                    formatted_date_str = formatted_date.strftime('%Y/%m')
                    
                    # print(year)
                    # month = yearmonthmatch.group(2)  # 月の部分を取得
                    # formatted_date = f"{year}/{month}"
                    print(f"Formatted date: {formatted_date_str}")
                else:
                    print(f"Date format doesn't match expected pattern.{filename}")
                    formatted_date = "noching"
                print(f"新品{new_price}　中古{used_price}")

                watch_item_instance.insert_data([make_ID,ref,year,model_name,size,bracelet,dial,url,"Watchnian(買取)",new_price,dateresult,otherdata,used_price,unused_price])
                print(make_ID)
                print("----------------")




def watchniansale(instance):
    current_directory = os.getcwd()
    folder_name = "まとめるエクセル"
    new_directory = os.path.join(current_directory, folder_name)
    items = instance.tablesdateill
    watch_item_instance = items["watch_item"]
    # ウォッチニアン販売
    pattern = re.compile(r'^ウォッチニアン.*\販売.*\.xlsx$')
    
    # 買取正規表現
    buygrammerpattern = r"(買取?)"
    salegrammerpattern = r"(販売)"
    numpattern = r"\d+"
    datepattern = r'(\d{4}).*(\d{2})'
    wachnian_files =[]
    for file in os.listdir(new_directory):
        if pattern.match(file):
            wachnian_files.append(os.path.join(new_directory, file))
    for file in wachnian_files:
        full_path = os.path.join(new_directory, file)
        workbook = load_workbook(full_path, data_only=False)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                rowitemslist = list(row)
                model_name = rowitemslist[0]
                ref = rowitemslist[1]
                year = ""
                size =""
                dial = rowitemslist[2]
                bracelet = rowitemslist[3]
                price = rowitemslist[4]
                new_price = 0
                used_price = 0
                unused_price = 0

                print(new_price)
                new_price = re.findall(numpattern,str(new_price))
                used_price = re.findall(numpattern,str(used_price))
                unused_price = re.findall(numpattern,str(unused_price))
                 
                try:
                    if isinstance(new_price[0],tuple):
                        new_price = "".join(new_price[0])
                except IndexError as e:
                    print("文字列or配列")
                except IndexError as e:
                    print("test")
                try:
                    if isinstance(used_price[0],tuple):
                        used_price = "".join(used_price[0])
                except IndexError as e:
                    print("文字列or配列")
                except IndexError as e:
                    print("test")
                try:
                    if isinstance(unused_price[0],tuple):
                        unused_price = "".join(unused_price[0])
                except IndexError as e:
                    print("文字列or配列")
                except IndexError as e:
                    print("test")
                
                if isinstance(new_price,list):
                    print(f"{new_price}はエラーになる値段")
                    new_price = "".join(new_price)
                    new_price = int(new_price)
                else:
                    new_price = 0
                
                if isinstance(used_price,list):
                    print(f"{unused_price}は中古の値段")
                    used_price = "".join(used_price)
                    used_price = int(used_price)
                else:
                    used_price = 0

                if isinstance(unused_price,list):
                    print(f"{unused_price}は未使用品の値段")
                    unused_price = "".join(unused_price)
                    unused_price = int(unused_price)
                else:
                    unused_price = 0
                url = rowitemslist[5]

                if not url:
                    url = rowitemslist[5]
                match = re.search(r'https://watchnian.com/shop/g/gik-.\d-+(\d+)', url)
                print(url)
                if match:
                    product_id = match.group(1)
                else:
                    print("----------------")
                    continue
               
                
         
                print(f"リファレンス:{ref} 年:{year} サイズ{size} ダイアル{dial} ブレスレット{bracelet} URL{url} 新品値段{new_price} 中古値段{used_price}")
                yearmonthpattern = re.compile(r'_(\d{4}-\d{2})')
                # ファイル名
                filename = os.path.basename(file)
                datematches = re.findall(datepattern,filename)
                print(datematches)
                dateresult = "/".join(datematches[0])
                make_ID_Format_Date = dateresult.replace("/","-")

                make_ID = "Watchnian-"+make_ID_Format_Date+"-"+product_id
                # 買取、販売チェックする。

                yearmonthmatch = re.search(r'_*(\d{4}-\d{2})', filename)
                # yearmonthmatch = yearmonthpattern.match(str(filename))
                buy_match = re.search(buygrammerpattern,filename)
                # 
                if buy_match:
                    otherdata = buy_match.group(1)
                else:
                    otherdata = "販売"
               
               
               
                if yearmonthmatch:
                    formatted_date = yearmonthmatch.group(1)  # 年月を取得する
                    formatted_date = datetime.strptime(formatted_date, '%Y-%m')
                    formatted_date_str = formatted_date.strftime('%Y/%m')
                    
                    # print(year)
                    # month = yearmonthmatch.group(2)  # 月の部分を取得
                    # formatted_date = f"{year}/{month}"
                    print(f"Formatted date: {formatted_date_str}")
                else:
                    print(f"Date format doesn't match expected pattern.{filename}")
                    formatted_date = "noching"
                print(f"新品{new_price}　中古{used_price}")

                watch_item_instance.insert_data([make_ID,ref,year,model_name,size,bracelet,dial,url,"Watchnian",price,dateresult,otherdata,used_price,unused_price])
                print(make_ID)
                print("----------------")

# 渡されたデータを日付に帰る関数
# ファイル名
def dateformatchenge(date):
        # 年
        yearmatch = re.search(r"\d{4}",date)
        # 年を削除する。
        yearromove = re.sub(r"(\d{4})","",date)
        # 月を取得する
        monthmach = re.search(r"(\d{1,2})",yearromove)
        yyyymm = []
        if yearmatch:
            extractionyear = yearmatch.group(0)
            yyyymm.append(extractionyear)
            if monthmach:
                month = monthmach.group(0)
                yyyymm.append(month)
                dateresult = "/".join(yyyymm)
                print(dateresult,"日付データ")
            else:
                print("日付がはいってない")
        return dateresult


# GMTのデータをエクセルからデータベースに入稿する処理
def gmt(instance):
    current_directory = os.getcwd()
    folder_name = "まとめるエクセル"
    new_directory = os.path.join(current_directory, folder_name)
    items = instance.tablesdateill
    watch_item_instance = items["watch_item"]
    # 年と月を抽出する正規表現パターン
    yearmonthpattern = re.compile(r'(\d{4})-(\d{2})')
    # 日付パターンを抽出する。
    datepattern = r"\d+"

    # フィールドを再定義する。
    # watch_item_instance.fields = ["item_id","ref","year","size","bracelet","dial","url"]
    # insert_values = [insert_date,"0",price,bucherer_watch_id]
    # print(insert_values)
    # watch_item_instance.insert_data(insert_values)
    # JWAファイル取得
    pattern = re.compile(r'^GMT.*\.xlsx$')
    buygrammerpattern = r"(買取り?)"
    jwa_files = []

    # ディレクトリ内のファイルを走査して、JWAファイルをリストに追加する
    for file in os.listdir(new_directory):
        if pattern.match(file):
            jwa_files.append(os.path.join(new_directory, file))
    for file in jwa_files:
        full_path = os.path.join(new_directory, file)
        workbook = load_workbook(full_path, data_only=False)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                print(list(row))
                rowitemslist = list(row)
                # 週フィールド名
                # weekly_table_fields = ['weekdate','ranking','price','bucherer_watch_id']
                # # 週レポート
                # weekly_report_inset_instance = SQLiteDataInsert(dbname,weekly_table_name,weekly_table_fields)
                # insert_values = [rowitemslist[1],"",rowitemslist[16],rowitemslist[10],"","",""]
                # insert_values = [rowitemslist[10],"",rowitemslist[16],"","","",""]
                # EVENCEのidはURLの一番下から抽出する。
                ref = rowitemslist[1]
                year = ""
                model_name = rowitemslist[0]
                size =""
                # いかみなおし！！
                dial = ""
                bracelet = rowitemslist[2]
                # 値段判定
                price = rowitemslist[4]
                print(price)
                price = re.findall(datepattern,price)
                print(price,"値段")
                if not price:
                    continue
                price = "".join(price)
                # if rowitemslist[3]:
                #     price = rowitemslist[3]
                # else:
                #     price = rowitemslist[4]
                url = rowitemslist[5]
                print(url)
                match = re.search(r'https://www.gmt-j.com/item/(\d+)', url)
                if match:
                    product_id = match.group(1)
                    print(product_id)
                else:
                    print("Product ID not found")
                    continue
                make_ID = "GMT-"+product_id
                print(make_ID)

                filename = os.path.basename(file)
                yearmonthmatch = yearmonthpattern.match(str(file))
                yearmonthmatch = re.search(r'(\d{4}-\d{2})', file)
                datematches = re.findall(datepattern,file)
                print(filename)
                # 年
                yearmatch = re.search(r"\d{4}",filename)
                # 年を削除する。
                yearromove = re.sub(r"(\d{4})","",filename)
                # 月を取得する
                monthmach = re.search(r"(\d{1,2})",yearromove)
                yyyymm = []
                if yearmatch:
                    extractionyear = yearmatch.group(0)
                    yyyymm.append(extractionyear)
                    if monthmach:
                        month = monthmach.group(0)
                        yyyymm.append(month)
                    dateresult = "/".join(yyyymm)
                    print(dateresult,"日付データ")
                else:
                    print("日付がはいってない")
                print(dateresult,"日付データ")
                dateresult = dateformatchenge(filename)
                # 日付
                # date_str = re.sub(r"\d{4}", "", datematches)
                # dateresult = "/".join(datematches)

                if yearmonthmatch:
                    formatted_date = yearmonthmatch.group(1)  # 年月を取得する
                    formatted_date = datetime.strptime(formatted_date, '%Y-%m')
                    formatted_date_str = formatted_date.strftime('%Y/%m')
                    
                    # print(year)
                    # month = yearmonthmatch.group(2)  # 月の部分を取得
                    # formatted_date = f"{year}/{month}"
                    print(f"Formatted date: {formatted_date_str}")
                else:
                    
                    print(f"Date format doesn't match expected pattern.{file}")
                    formatted_date = "noching"
                
                # 品番?
                # if rowitemslist[10] is None:
                #     # rowitemslist[10] が None の場合の処理
                #     print("rowitemslist[10] は None です")
                #     ref = "なし"
                # else:
                #      # rowitemslist[10] が None でない場合の処理
                #      print("rowitemslist[10] は None ではありません")
                #      ref = rowitemslist[10] 
                    
                # ↓あとで使うやつ
                # ↓その他アイテム
                
                watch_item_instance.insert_data([make_ID,ref,year,model_name,size,bracelet,dial,url,"GMT",price,dateresult,"その他","",""])

                # watch_item_instance.fieldcountAllcountcheck()
                row_data = []
                # dbにデータを入れる
                # for cell in row: 
                #     print(str(cell))      
            # insert_values = [insert_date,"0",price,bucherer_watch_id]
    return jwa_files

# なんぼや(買取)
def nanboya(instance):
    current_directory = os.getcwd()
    folder_name = "まとめるエクセル"
    new_directory = os.path.join(current_directory, folder_name)
    items = instance.tablesdateill
    watch_item_instance = items["watch_item"]
    # 年と月を抽出する正規表現パターン
    yearmonthpattern = re.compile(r'(\d{4})-(\d{2})')
    # 日付パターンを抽出する。
    datepattern = r"\d+"

    # フィールドを再定義する。
    # watch_item_instance.fields = ["item_id","ref","year","size","bracelet","dial","url"]
    # insert_values = [insert_date,"0",price,bucherer_watch_id]
    # print(insert_values)
    # watch_item_instance.insert_data(insert_values)
    # JWAファイル取得
    pattern = re.compile(r'^なんぼや.*\.xlsx$')
    buygrammerpattern = r"(買取り?)"
    jwa_files = []
   
    testnum =0
    # ディレクトリ内のファイルを走査して、JWAファイルをリストに追加する
    for file in os.listdir(new_directory):
        if pattern.match(file):
            jwa_files.append(os.path.join(new_directory, file))
    print(len(jwa_files))
    for file in jwa_files:
        full_path = os.path.join(new_directory, file)
        workbook = load_workbook(full_path, data_only=False)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                print(list(row))
                rowitemslist = list(row)
                # 週フィールド名
                # weekly_table_fields = ['weekdate','ranking','price','bucherer_watch_id']
                # # 週レポート
                # weekly_report_inset_instance = SQLiteDataInsert(dbname,weekly_table_name,weekly_table_fields)
                # insert_values = [rowitemslist[1],"",rowitemslist[16],rowitemslist[10],"","",""]
                # insert_values = [rowitemslist[10],"",rowitemslist[16],"","","",""]
                # EVENCEのidはURLの一番下から抽出する。
                ref = rowitemslist[1]
                if not ref:
                    ref = "NotReference"
                year = ""
                model_name = rowitemslist[0]
                size =""
                dial = rowitemslist[2]
                bracelet = ""
                # 値段判定
                price = rowitemslist[3]
                print(price)
                price = re.findall(datepattern,str(price))
                print(price,"値段")
                if not price:
                    continue
                price = "".join(price)
                price = int(price)
                # if rowitemslist[3]:
                #     price = rowitemslist[3]
                # else:
                #     price = rowitemslist[4]
                url = rowitemslist[4]
                match = re.search(r'.*/([^/]+)/?$', url)
                # 最後/にあるパターンがもれる
                match = re.search(r'.*/([^/]+?)(?:\.html)?/?$', url)
                if match:
                    product_id = match.group(1)
                    testnum += 1
                    print(testnum)
                    print(product_id)

                else:
                    print(f"Product ID not found{url}")
                    print("------------------------------")
                    continue
                make_ID = "nanboya"+product_id
                print(make_ID)

                filename = os.path.basename(file)
                yearmonthmatch = yearmonthpattern.match(str(file))
                yearmonthmatch = re.search(r'(\d{4}-\d{2})', file)
                datematches = re.findall(datepattern,file)
                print(filename)
                # 年
                yearmatch = re.search(r"\d{4}",filename)
                # 年を削除する。
                yearromove = re.sub(r"(\d{4})","",filename)
                # 月を取得する
                monthmach = re.search(r"(\d{1,2})",yearromove)
                yyyymm = []
                if yearmatch:
                    extractionyear = yearmatch.group(0)
                    yyyymm.append(extractionyear)
                    if monthmach:
                        month = monthmach.group(0)
                        yyyymm.append(month)
                    dateresult = "/".join(yyyymm)
                    print(dateresult,"日付データ")
                else:
                    print("日付がはいってない")
                print(dateresult,"日付データ")
                dateresult = dateformatchenge(filename)
                # 日付
                # date_str = re.sub(r"\d{4}", "", datematches)
                # dateresult = "/".join(datematches)
                buy_match = re.search(buygrammerpattern,filename)
                # 
                if buy_match:
                    otherdata = buy_match.group(1)
                else:
                    otherdata = "販売"
            
                
                # 品番?
                # if rowitemslist[10] is None:
                #     # rowitemslist[10] が None の場合の処理
                #     print("rowitemslist[10] は None です")
                #     ref = "なし"
                # else:
                #      # rowitemslist[10] が None でない場合の処理
                #      print("rowitemslist[10] は None ではありません")
                #      ref = rowitemslist[10] 
                    
                # ↓あとで使うやつ
                # ↓その他アイテム
                
                watch_item_instance.insert_data([make_ID,ref,year,model_name,size,bracelet,dial,url,"なんぼや(買取)",price,dateresult,otherdata,"",""])

                # watch_item_instance.fieldcountAllcountcheck()
                row_data = []
                # dbにデータを入れる
                # for cell in row: 
                #     print(str(cell))      
            # insert_values = [insert_date,"0",price,bucherer_watch_id]
    return jwa_files



# クオーク
def quark(instance):
    current_directory = os.getcwd()
    folder_name = "まとめるエクセル"
    new_directory = os.path.join(current_directory, folder_name)
    items = instance.tablesdateill
    watch_item_instance = items["watch_item"]
    # 年と月を抽出する正規表現パターン
    yearmonthpattern = re.compile(r'(\d{4})-(\d{2})')
    # 日付パターンを抽出する。
    datepattern = r"\d+"

    # フィールドを再定義する。
    # watch_item_instance.fields = ["item_id","ref","year","size","bracelet","dial","url"]
    # insert_values = [insert_date,"0",price,bucherer_watch_id]
    # print(insert_values)
    # watch_item_instance.insert_data(insert_values)
    # JWAファイル取得
    pattern = re.compile(r'^クオーク.*\.xlsx$')
    buygrammerpattern = r"(買取り?)"
    jwa_files = []
    print("クオーク関数")
    testnum =0
    # ディレクトリ内のファイルを走査して、JWAファイルをリストに追加する
    for file in os.listdir(new_directory):
        if pattern.match(file):
            jwa_files.append(os.path.join(new_directory, file))
    print(len(jwa_files))
    for file in jwa_files:
        full_path = os.path.join(new_directory, file)
        workbook = load_workbook(full_path, data_only=False)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                print(list(row))
                rowitemslist = list(row)
                # 週フィールド名
                # weekly_table_fields = ['weekdate','ranking','price','bucherer_watch_id']
                # # 週レポート
                # weekly_report_inset_instance = SQLiteDataInsert(dbname,weekly_table_name,weekly_table_fields)
                # insert_values = [rowitemslist[1],"",rowitemslist[16],rowitemslist[10],"","",""]
                # insert_values = [rowitemslist[10],"",rowitemslist[16],"","","",""]
                # EVENCEのidはURLの一番下から抽出する。
                ref = rowitemslist[1]
                year = ""
                model_name = rowitemslist[0]
                size =""
                dial = rowitemslist[2]
                bracelet = rowitemslist[3]
                # 値段判定
                price = rowitemslist[4]
                print(price)
                price = re.findall(datepattern,str(price))
                print(price,"値段")
                if not price:
                    continue
                price = "".join(price)
                price = int(price)
                # if rowitemslist[3]:
                #     price = rowitemslist[3]
                # else:
                #     price = rowitemslist[4]
                url = rowitemslist[5]
                print(url)
                # match = re.search(r'https://www.909.co.jp/rolex_catalog/.*?(\d+).*?/', url)
                match = re.search(r'https://www.909.co.jp/rolex_catalog/.*?(\d+).*?/', url)

                match = re.search(r'.*/([^/]+)/?$', url)
                # 最後/にあるパターンがもれる
                match = re.search(r'.*/([^/]+?)(?:\.html)?/?$', url)

                if match:
                    product_id = match.group(1)
                    testnum += 1
                    print(testnum)
                    print(product_id)

                else:
                    print(f"Product ID not found{url}")
                    print("------------------------------")
                    continue
                

                filename = os.path.basename(file)
                yearmonthmatch = yearmonthpattern.match(str(file))
                yearmonthmatch = re.search(r'(\d{4}-\d{2})', file)
                datematches = re.findall(datepattern,file)
                print(filename)
                # 年
                yearmatch = re.search(r"\d{4}",filename)
                # 年を削除する。
                yearromove = re.sub(r"(\d{4})","",filename)
                # 月を取得する
                monthmach = re.search(r"(\d{1,2})",yearromove)
                yyyymm = []
                if yearmatch:
                    extractionyear = yearmatch.group(0)
                    yyyymm.append(extractionyear)
                    if monthmach:
                        month = monthmach.group(0)
                        yyyymm.append(month)
                    dateresult = "/".join(yyyymm)
                    print(dateresult,"日付データ")
                else:
                    print("日付がはいってない")
                print(dateresult,"日付データ")
                dateresult = dateformatchenge(filename)
                # 日付
                # date_str = re.sub(r"\d{4}", "", datematches)
                # dateresult = "/".join(datematches)

                if yearmonthmatch:
                    formatted_date = yearmonthmatch.group(1)  # 年月を取得する
                    formatted_date = datetime.strptime(formatted_date, '%Y-%m')
                    formatted_date_str = formatted_date.strftime('%Y/%m')
                    
                    # print(year)
                    # month = yearmonthmatch.group(2)  # 月の部分を取得
                    # formatted_date = f"{year}/{month}"
                    print(f"Formatted date: {formatted_date_str}")
                else:
                    
                    print(f"Date format doesn't match expected pattern.{file}")
                    formatted_date = "noching"
                
                make_ID_Format_Date = dateresult.replace("/","-")
                make_ID = "quark-"+make_ID_Format_Date+"-"+product_id
                # 品番?
                # if rowitemslist[10] is None:
                #     # rowitemslist[10] が None の場合の処理
                #     print("rowitemslist[10] は None です")
                #     ref = "なし"
                # else:
                #      # rowitemslist[10] が None でない場合の処理
                #      print("rowitemslist[10] は None ではありません")
                #      ref = rowitemslist[10] 
                    
                # ↓あとで使うやつ
                # ↓その他アイテム
                
                watch_item_instance.insert_data([make_ID,ref,year,model_name,size,bracelet,dial,url,"クオーク",price,dateresult,"その他","",""])

                # watch_item_instance.fieldcountAllcountcheck()
                row_data = []
                # dbにデータを入れる
                # for cell in row: 
                #     print(str(cell))      
            # insert_values = [insert_date,"0",price,bucherer_watch_id]
    return jwa_files

# 価格コム
def kakakucom(instance):
    current_directory = os.getcwd()
    folder_name = "まとめるエクセル"
    new_directory = os.path.join(current_directory, folder_name)
    items = instance.tablesdateill
    watch_item_instance = items["watch_item"]
    # 年と月を抽出する正規表現パターン
    yearmonthpattern = re.compile(r'(\d{4})-(\d{2})')
    # 日付パターンを抽出する。
    datepattern = r"\d+"

    # フィールドを再定義する。
    # watch_item_instance.fields = ["item_id","ref","year","size","bracelet","dial","url"]
    # insert_values = [insert_date,"0",price,bucherer_watch_id]
    # print(insert_values)
    # watch_item_instance.insert_data(insert_values)
    # JWAファイル取得
    pattern = re.compile(r'^価格.*\.xlsx$')
    buygrammerpattern = r"(買取り?)"
    jwa_files = []
    print("クオーク関数")
    testnum =0
    # ディレクトリ内のファイルを走査して、JWAファイルをリストに追加する
    for file in os.listdir(new_directory):
        if pattern.match(file):
            jwa_files.append(os.path.join(new_directory, file))
    print(len(jwa_files))
    for file in jwa_files:
        full_path = os.path.join(new_directory, file)
        workbook = load_workbook(full_path, data_only=False)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                print(list(row))
                rowitemslist = list(row)
                # 週フィールド名
                # weekly_table_fields = ['weekdate','ranking','price','bucherer_watch_id']
                # # 週レポート
                # weekly_report_inset_instance = SQLiteDataInsert(dbname,weekly_table_name,weekly_table_fields)
                # insert_values = [rowitemslist[1],"",rowitemslist[16],rowitemslist[10],"","",""]
                # insert_values = [rowitemslist[10],"",rowitemslist[16],"","","",""]
                # EVENCEのidはURLの一番下から抽出する。
                ref = rowitemslist[1]
                year = ""
                model_name = rowitemslist[0]
                size =""
                #ブレスレット
                bracelet = rowitemslist[2]
                dial = ""

                # 新品値段
                new_price = rowitemslist[3]
                print(new_price,"新品値段")
                new_price = re.findall(datepattern,str(new_price))
                if new_price:
                    new_price = re.findall(datepattern,str(new_price))
                    new_price = new_price[0]
                    print(new_price)
                    print(new_price[0],"新品の値段")
                    new_price = new_price
                else:
                    new_price = 0
                
                
                # 中古値段
                use_price = rowitemslist[4]
                use_price = re.findall(datepattern,str(use_price))
                if use_price:
                    
                    use_price = use_price[0]
                else:
                    use_price = 0
                
                
                # price = "".join(price)
                # price = int(price)
            
                ranking = rowitemslist[5]
                url = rowitemslist[6]
                print(url)
                match = re.search(r'.*/([^/]+)/?$', url)
                # 最後/にあるパターンがもれる
                match = re.search(r'.*/([^/]+?)(?:\.html)?/?$', url)

                if match:
                    product_id = match.group(1)
                    testnum += 1
                    print(testnum)
                    print(product_id)

                else:
                    print(f"Product ID not found{url}")
                    print("------------------------------")
                    continue
                make_ID = "kakakucom-"+product_id
                filename = os.path.basename(file)
                yearmonthmatch = yearmonthpattern.match(str(file))
                yearmonthmatch = re.search(r'(\d{4}-\d{2})', file)
                datematches = re.findall(datepattern,file)
                print(filename)
                # 年
                yearmatch = re.search(r"\d{4}",filename)
                # 年を削除する。
                yearromove = re.sub(r"(\d{4})","",filename)
                # 月を取得する
                monthmach = re.search(r"(\d{1,2})",yearromove)
                yyyymm = []
                if yearmatch:
                    extractionyear = yearmatch.group(0)
                    yyyymm.append(extractionyear)
                    if monthmach:
                        month = monthmach.group(0)
                        yyyymm.append(month)
                    dateresult = "/".join(yyyymm)
                    print(dateresult,"日付データ")
                else:
                    print("日付がはいってない")
                print(dateresult,"日付データ")
                dateresult = dateformatchenge(filename)
                # 日付
                # date_str = re.sub(r"\d{4}", "", datematches)
                # dateresult = "/".join(datematches)

                # 品番?
                # if rowitemslist[10] is None:
                #     # rowitemslist[10] が None の場合の処理
                #     print("rowitemslist[10] は None です")
                #     ref = "なし"
                # else:
                #      # rowitemslist[10] が None でない場合の処理
                #      print("rowitemslist[10] は None ではありません")
                #      ref = rowitemslist[10] 
                    
                # ↓あとで使うやつ
                # ↓その他アイテム

                # データ挿入
                watch_item_instance.insert_data([make_ID,ref,year,model_name,size,bracelet,dial,url,"価格コム",new_price,dateresult,ranking,use_price,""])

                # watch_item_instance.fieldcountAllcountcheck()
                row_data = []
                # dbにデータを入れる
                # for cell in row: 
                #     print(str(cell))      
            # insert_values = [insert_date,"0",price,bucherer_watch_id]
    return jwa_files


def evence(instance):
    current_directory = os.getcwd()
    folder_name = "まとめるエクセル"
    new_directory = os.path.join(current_directory, folder_name)
    items = instance.tablesdateill
    watch_item_instance = items["watch_item"]
    # 年と月を抽出する正規表現パターン
    yearmonthpattern = re.compile(r'(\d{4})-(\d{2})')
    # フィールドを再定義する。
    # watch_item_instance.fields = ["item_id","ref","year","size","bracelet","dial","url"]
    # insert_values = [insert_date,"0",price,bucherer_watch_id]
    # print(insert_values)
    # watch_item_instance.insert_data(insert_values)
    # JWAファイル取得
    pattern = re.compile(r'^EVANCE.*\.xlsx$')
    jwa_files = []

    # ディレクトリ内のファイルを走査して、JWAファイルをリストに追加する
    for file in os.listdir(new_directory):
        if pattern.match(file):
            jwa_files.append(os.path.join(new_directory, file))
    for file in jwa_files:
        full_path = os.path.join(new_directory, file)
        workbook = load_workbook(full_path, data_only=False)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                print(list(row))
                rowitemslist = list(row)
                # 週フィールド名
                # weekly_table_fields = ['weekdate','ranking','price','bucherer_watch_id']
                # # 週レポート
                # weekly_report_inset_instance = SQLiteDataInsert(dbname,weekly_table_name,weekly_table_fields)
                # insert_values = [rowitemslist[1],"",rowitemslist[16],rowitemslist[10],"","",""]
                # insert_values = [rowitemslist[10],"",rowitemslist[16],"","","",""]
                # EVENCEのidはURLの一番下から抽出する。
                ref = rowitemslist[1]
                year = ""
                model_name = rowitemslist[0]
                size =""
                dial = ""
                bracelet = rowitemslist[2]
                # 値段判定
                if rowitemslist[3]:
                    price = rowitemslist[3]
                else:
                    price = 0
                # 中古値段判定
                if rowitemslist[4]:
                    used_price = rowitemslist[4]
                else:
                    used_price = 0
                url = rowitemslist[5]
                print(url)
                match = re.search(r'/products/detail/(\d+)', url)
                if match:
                    product_id = match.group(1)
                    print(product_id)
                else:
                    print("Product ID not found")
                    continue
                make_ID = "Evence-"+product_id
                print(make_ID)




                yearmonthmatch = yearmonthpattern.match(str(file))
                yearmonthmatch = re.search(r'(\d{4}-\d{2})', file)
                if yearmonthmatch:
                    formatted_date = yearmonthmatch.group(1)  # 年月を取得する
                    formatted_date = datetime.strptime(formatted_date, '%Y-%m')
                    formatted_date_str = formatted_date.strftime('%Y/%m')
                    
                    # print(year)
                    # month = yearmonthmatch.group(2)  # 月の部分を取得
                    # formatted_date = f"{year}/{month}"
                    print(f"Formatted date: {formatted_date_str}")
                else:
                    
                    print(f"Date format doesn't match expected pattern.{file}")
                    formatted_date_str = "noching"
                
                # 品番?
                # if rowitemslist[10] is None:
                #     # rowitemslist[10] が None の場合の処理
                #     print("rowitemslist[10] は None です")
                #     ref = "なし"
                # else:
                #      # rowitemslist[10] が None でない場合の処理
                #      print("rowitemslist[10] は None ではありません")
                #      ref = rowitemslist[10] 
                    
                # ↓あとで使うやつ
                # ↓その他アイテム
                print(formatted_date_str,"値確認")
                watch_item_instance.insert_data([make_ID,ref,year,model_name,size,bracelet,dial,url,"EVENCE",price,formatted_date_str,"その他",used_price,""])

                # watch_item_instance.fieldcountAllcountcheck()
                row_data = []
                # dbにデータを入れる
                # for cell in row: 
                #     print(str(cell))      
            # insert_values = [insert_date,"0",price,bucherer_watch_id]
    return jwa_files



# ブヘラ(USA)のデータを統合データベースに入稿する処理

# jwaのデータをデータベースに入稿する処理
def jwa(instance):
    current_directory = os.getcwd()
    folder_name = "まとめるエクセル"
    new_directory = os.path.join(current_directory, folder_name)
    items = instance.tablesdateill
    watch_item_instance = items["watch_item"]
    # 年と月を抽出する正規表現パターン
    yearmonthpattern = re.compile(r'(\d{4})-(\d{2})')
    # フィールドを再定義する。
    # watch_item_instance.fields = ["item_id","ref","year","size","bracelet","dial","url"]
    # insert_values = [insert_date,"0",price,bucherer_watch_id]
    # print(insert_values)
    # watch_item_instance.insert_data(insert_values)
    # JWAファイル取得
    pattern = re.compile(r'^JWA.*\.xlsx$')
    jwa_files = []

    # ディレクトリ内のファイルを走査して、JWAファイルをリストに追加する
    for file in os.listdir(new_directory):
        if pattern.match(file):
            jwa_files.append(os.path.join(new_directory, file))
    # print(jwa_files)
    for file in jwa_files:
        full_path = os.path.join(new_directory, file)
        workbook = load_workbook(full_path, data_only=False)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                print(list(row))
                rowitemslist = list(row)
                # 週フィールド名
                # weekly_table_fields = ['weekdate','ranking','price','bucherer_watch_id']
                # # 週レポート
                # weekly_report_inset_instance = SQLiteDataInsert(dbname,weekly_table_name,weekly_table_fields)
                insert_values = [rowitemslist[1],"",rowitemslist[16],rowitemslist[10],"","",""]
                insert_values = [rowitemslist[10],"",rowitemslist[16],"","","",""]
                # 
                jwa_make_ID = "JWA-"+str(rowitemslist[1])+"-"+str(rowitemslist[2])+"-"+str(rowitemslist[3])
                yearmonthmatch = yearmonthpattern.match(str(rowitemslist[1]))

                if yearmonthmatch:
                    year = yearmonthmatch.group(1)  # 年の部分を取得
                    month = yearmonthmatch.group(2)  # 月の部分を取得
                    formatted_date = f"{year}/{month}"
                    print(f"Formatted date: {formatted_date}")
                else:
                    print("Date format doesn't match expected pattern.")
                    formatted_date = "noching"
                
                # 品番?
                if rowitemslist[10] is None:
                    # rowitemslist[10] が None の場合の処理
                    print("rowitemslist[10] は None です")
                    ref = "なし"
                else:
                     # rowitemslist[10] が None でない場合の処理
                     print("rowitemslist[10] は None ではありません")
                     ref = rowitemslist[10] 
                    
                # ↓あとで使うやつ
                # ↓その他アイテム
                
                watch_item_instance.insert_data([jwa_make_ID,ref,"","","","","","","JWA",rowitemslist[16],formatted_date,rowitemslist[11],""])

                # watch_item_instance.fieldcountAllcountcheck()
                row_data = []
                # dbにデータを入れる
                # for cell in row: 
                #     print(str(cell))      
            # insert_values = [insert_date,"0",price,bucherer_watch_id]
    return jwa_files

def refdatasget(instance):
    items = instance.tablesdateill
    watch_item_instance = items["watch_item"]
    returnitems = watch_item_instance.groupby('test')
    return returnitems

# JBA(買取)
def jba(instance):
    current_directory = os.getcwd()
    folder_name = "まとめるエクセル"
    new_directory = os.path.join(current_directory, folder_name)
    items = instance.tablesdateill
    watch_item_instance = items["watch_item"]
    # 年と月を抽出する正規表現パターン
    yearmonthpattern = re.compile(r'(\d{4})-(\d{2})')
    # 日付パターンを抽出する。
    datepattern = r"\d+"

    # フィールドを再定義する。
    # watch_item_instance.fields = ["item_id","ref","year","size","bracelet","dial","url"]
    # insert_values = [insert_date,"0",price,bucherer_watch_id]
    # print(insert_values)
    # watch_item_instance.insert_data(insert_values)
    # JWAファイル取得
    pattern = re.compile(r'^jba.*\.xlsx$')
    pattern = re.compile(r'^jba.*\.xlsx$', re.IGNORECASE)
    buygrammerpattern = r"(買取り?)"
    jwa_files = []
   
    testnum = 0
    # ディレクトリ内のファイルを走査して、JBAファイルをリストに追加する
    for file in os.listdir(new_directory):
        if pattern.match(file):
            jwa_files.append(os.path.join(new_directory, file))
    print(len(jwa_files))
    for file in jwa_files:
        full_path = os.path.join(new_directory, file)
        workbook = load_workbook(full_path, data_only=False)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                print(list(row))
                rowitemslist = list(row)
                all_none = all(x is None for x in rowitemslist)
                if all_none:
                    continue
                # 週フィールド名
                # weekly_table_fields = ['weekdate','ranking','price','bucherer_watch_id']
                # # 週レポート
                # weekly_report_inset_instance = SQLiteDataInsert(dbname,weekly_table_name,weekly_table_fields)
                # insert_values = [rowitemslist[1],"",rowitemslist[16],rowitemslist[10],"","",""]
                # insert_values = [rowitemslist[10],"",rowitemslist[16],"","","",""]
                # EVENCEのidはURLの一番下から抽出する。
                ref = rowitemslist[10]
                column_num = 0
                if not ref:
                    ref = "NotReference"
                    column_num = 1
                year = ""
                model_name = rowitemslist[0]
                size =""
                dial = ""
                bracelet = ""
                # 値段判定
                print(f"{column_num}は引く数")
                price = rowitemslist[16 - column_num]
                print(price)
                price = re.findall(datepattern,str(price))
                print(price,"値段")
                if not price:
                    price = 0
                else:
                    price = "".join(price)
                    price = int(price)
                date = rowitemslist[1]
                # 中古の値段をいれる
                # datetimeオブジェクトに変換
                try:
                    # date_obj = datetime.strptime(date, "%Y/%m")
                    date_obj = datetime.strptime(date, '%Y/%m/%d')
                    formatted_date = date_obj.strftime("%Y-%m-%d")
                except ValueError as e:
                    print(date,"は日付でないといけない")
                    print("エラーに入ってる")
                    continue
                # "yyyy-mm" 形式に変換
                # formatted_date = date_obj.strftime("%Y-%m")
                
                # するのはどちらが難題か
                # if rowitemslist[3]:
                #     price = rowitemslist[3]
                # else:
                #     price = rowitemslist[4]
                url = rowitemslist[4]
                match = re.search(r'.*/([^/]+)/?$', url)
                # 最後/にあるパターンがもれる
                match = re.search(r'.*/([^/]+?)(?:\.html)?/?$', url)
                if match:
                    product_id = match.group(1)
                    testnum += 1
                    print(testnum)
                    print(product_id)

                else:
                    print(f"Product ID not found{url}")
                    print("------------------------------")
                    url = ""
                    # continue
               
 

                filename = os.path.basename(file)
                yearmonthmatch = yearmonthpattern.match(str(file))
                yearmonthmatch = re.search(r'(\d{4}-\d{2})', file)
                datematches = re.findall(datepattern,file)
                print(filename)
                # 年
                yearmatch = re.search(r"\d{4}",filename)
                # 年を削除する。
                yearromove = re.sub(r"(\d{4})","",filename)
                # 月を取得する
                monthmach = re.search(r"(\d{1,2})",yearromove)
                yyyymm = []
                if yearmatch:
                    extractionyear = yearmatch.group(0)
                    yyyymm.append(extractionyear)
                    if monthmach:
                        month = monthmach.group(0)
                        yyyymm.append(month)
                    dateresult = "/".join(yyyymm)
                    print(dateresult,"日付データ")
                else:
                    print("日付がはいってない")
                print(date,"日付データ")
                dateresult = dateformatchenge(date)
                print(formatted_date)
                # dateresult = date
                make_ID = "JBA-"+formatted_date+"-"+str(rowitemslist[2])+"-"+str(rowitemslist[3])
                print(f"{make_ID}はつくられたIDです")
                # 日付
                # date_str = re.sub(r"\d{4}", "", datematches)
                # dateresult = "/".join(datematches)
                buy_match = re.search(buygrammerpattern,filename)
                # 
                if buy_match:
                    otherdata = buy_match.group(1)
                else:
                    otherdata = "販売"
            
                otherdata = rowitemslist[11]



                # 品番?
                # if rowitemslist[10] is None:
                #     # rowitemslist[10] が None の場合の処理
                #     print("rowitemslist[10] は None です")
                #     ref = "なし"
                # else:
                #      # rowitemslist[10] が None でない場合の処理
                #      print("rowitemslist[10] は None ではありません")
                #      ref = rowitemslist[10] 
                    
                # ↓あとで使うやつ
                # ↓その他アイテム
          
                watch_item_instance.insert_data([make_ID,ref,year,model_name,size,bracelet,dial,url,"JBA(買取)",price,dateresult,otherdata,"",""])

                # watch_item_instance.fieldcountAllcountcheck()
                row_data = []
                # dbにデータを入れる
                # for cell in row: 
                #     print(str(cell))      
            # insert_values = [insert_date,"0",price,bucherer_watch_id]
    return jwa_files



def main():
    today_date = datetime.now().date()
    # エクセルファイル作成
    file_name = f"総合リスト{today_date}.xlsx"

    if not os.path.exists(file_name):
        # Excelブックの作成
        wb = Workbook()
        ws = wb.active
        # ヘッダー行を追加
        ws.append(
            [
             "リファレンス",
             "年代",
             "モデル名",
             "サイズ",
             "ブレスレット",
             "ダイアル",
             "値段"
             ]
    )     
    else:
        # ファイルが存在する場合は既存のファイルを読み込み
        wb = load_workbook(file_name)
        ws = wb.active

    # インスタンスを作成する。
    db_file = "bucherer.db"
    # dbのファイルをわたす。
    whocequeryinstance = WhocheSqliteDataInsert(db_file)
    # jba(whocequeryinstance)
    # return
    # nanboya(whocequeryinstance)
    # # return

    # kakakucom(whocequeryinstance)
    # # return

    # quark(whocequeryinstance)
    # return

    # 月曜日は以下を抽出!!
    # gmt(whocequeryinstance)
    # return
    # wachnianbuy

    # watchniansale(whocequeryinstance)
    # return
    # watchnianbuy(whocequeryinstance)
    # return

    # evence(whocequeryinstance)
    # return

    # jwa(whocequeryinstance)
    # return
    # ここで列を指定する。
    excel_alldata = refdatasget(whocequeryinstance) 
    for data in excel_alldata:
        # print(data,"データ一覧")
        ws.append(data)
        # wb.save(file_name)
    wb.save(file_name)

    return
    
    today_date_and_time = datetime.now().strftime("%Y%m%d%H%M%S")
    today_date = datetime.now().strftime("%Y%m%d")
    file_name = f"データ集{today_date}.xlsx"  
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
    else:
        wb = load_workbook(file_name)
        ws = wb.active

    current_directory = os.getcwd()
    folder_name = "まとめるエクセル"
    dest_folder = os.path.join(current_directory, f"{folder_name}copy")
    new_directory = os.path.join(current_directory, folder_name)
    
    copy_folder(new_directory, dest_folder)
    excel_files = [file for file in os.listdir(new_directory) if file.endswith('.xlsx') or file.endswith('.xls')]

    Ref_pattern = re.compile(r"\b\d{4,6}[A-Za-z]*\b")

    for file in excel_files:
        full_path = os.path.join(new_directory, file)
        workbook = load_workbook(full_path, data_only=False)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_data = []
                for cell in row:               
                    if cell is not None:
                        Ref_matches = Ref_pattern.findall(str(cell))
                        if Ref_matches:
                            print(Ref_matches,"Ref")
                    row_data.append(cell)
                ws.append(row_data) 
        
        wb.save(file_name)  
        print("Sheets:", wb.sheetnames)

if __name__ == "__main__":
    main()
