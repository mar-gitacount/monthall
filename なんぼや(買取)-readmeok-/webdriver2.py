from selenium import webdriver
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
import os
import sys
from selenium.webdriver.chrome.options import Options
num = int(sys.argv[1]) if len(sys.argv) > 1 else 0

# ここで num を使用して何かを実行する
(f"受け取った引数: {num}")
url = "https://kakaku.com/watch_accessory/watch/itemlist.aspx?pdf_ma=5090&pdf_Spec106=1,2&pdf_vi=c"
url = f"https://nanboya.com/search/item-list/b-948/?page={num}"


# ?モデル、ブレスレット、文字盤を抽出する関数
def model_validete_imput(text):
    models = [
        "デイトジャスト",
        "オイスター",
        "コスモグラフ",
        "シードゥエラー",
        "エクスプローラー",
        "GMTマスター",
        "GMTマスターII",
        "サブマリーナー",
        "ヨットマスター",
        "スカイドゥエラー",
        "エクスプローラーII",
        "エアキング",
    ]
    print("関数内のテキスト", text)
    pattern = r"\b\s+(\S+)\s+\b"
    beltpattern = r"\[(.*?)\]|\((.*?)\)"
    # ベルトと文字盤を抽出する。
    beltmatches = re.findall(beltpattern, text)
    # モデル名を抽出する。
    model = re.sub(beltpattern, "", text)
    print("モデル名", model)
    # [],()を正規表現で抽出する。
    items = {"model": model, "beltmatches": beltmatches}
    return items


# ?エクセルに入力する関数
def wsinsert(values, sheet):
    print("wsinsert関数", values)
    sheet.append(values)
    # for item in values:
    #     sheet.append(item)


# 現在の日付を取得
today_date = datetime.now().strftime("%Y%m%d")
# ファイル名に日付を組み込む

file_name = f"なんぼや_{today_date}買取.xlsx"
if not os.path.exists(file_name):
    # Excelブックの作成
    wb = Workbook()
    ws = wb.active
    # ヘッダー行を追加
    ws.append(
        [
            "モデル名",
            "リファレンスNO",
            "文字盤",
            "買取価格",
            "URL",
        ]
    )
else:
    # ファイルが存在する場合は既存のファイルを読み込み
    wb = load_workbook(file_name)
    ws = wb.active


# SeleniumのWebDriverを初期化
options = Options()
options.add_argument("--headless")  # ヘッドレスモードを有効にする
driver = webdriver.Chrome(options=options)  # または他のブラウザに合わせて選択

# URLを開く
driver.get(url)

# Seleniumがページのロードを待つなどの適切な待機処理が必要な場合はここで実施

# ページのHTMLを取得
page_source = driver.page_source

# BeautifulSoupを使ってHTMLを解析
soup = BeautifulSoup(page_source, "html.parser")


# !ここから処理スタート
#?入稿可能な変数
# <tbody> タグ内のテキストを抽出して表示
tbody_tag = soup.find("body")
main_tag = tbody_tag.find("main", class_="main")
items = main_tag.find("section", class_="purchase-archive__box")
cards = items.find_all("div", class_="archive-cards")
for item in cards:
    # 各アイテムの変数を初期化
    model = ""
    ref = ""
    dial = ""
    # insertitems = []
    a_tag_get = item.find("a", class_="archive-cards__link")
    # 各アイテムごとのURLなので、これをもとにアクセスしてデータを抽出する
    href = a_tag_get.get("href")
    driver.get(href)

    items_page_source = driver.page_source
    item_soup = BeautifulSoup(items_page_source, "html.parser")
    
    # 各アイテムのbodyタグ
    cards_tbody_tag = item_soup.find("body")
    # 各アイテムのmainタグ
    cards_main_tag = cards_tbody_tag.find("main", class_="main")
    # print(cards_main_tag)
    cards_purchase_tag = cards_main_tag.find("div", class_="purchase")
    # ?金額。
    cards_price_text = cards_purchase_tag.find("p",class_="purchase-price__price-box").find("span",class_="purchase-price__price-number").get_text(strip=True)
    # 素材等を取得する。
    cards_purchase_table_tbody_tag = cards_purchase_tag.find("table", class_="purchase-item__table").find("tbody")

    # thタグを取得してループ各時計の要素がある。
    cards_purchase_th_tag = cards_purchase_table_tbody_tag.find_all("th")

    for th_tag in cards_purchase_th_tag:
        th_tag_text = th_tag.get_text(strip=True)
        print(th_tag_text)
        if th_tag_text == "モデル":
            td_tag_text = th_tag.find_next_sibling("td").get_text(strip=True)
            model = td_tag_text
            # tdタグの値を抽出する。
            print(model)
        elif th_tag_text == "型番":
            ref = td_tag_text = th_tag.find_next_sibling("td").get_text(strip=True)
            print(ref)
        elif th_tag_text == "文字盤":
            dial = th_tag.find_next_sibling("td").get_text(strip=True)
            print(dial)
    # ?モデルが空ならば、時計以外のアイテムなので、continue
    if model == "":
        continue
    # ?ここで、アイテムごとの配列を入れ込む。
    insertitems=[model,ref,dial,cards_price_text,href]
    wsinsert(insertitems,ws)
 
    insertitems = []
    
    
    print("--------------------------")
wb.save(file_name)
