import subprocess
import sys
import moduleuse

try:
    # ここに通常のコードを書く、例えば:
    # import numpy
    # import pandas
    # といった具体的なインポート文を記述する
    from selenium import webdriver
    import os
    import json
    from bs4 import BeautifulSoup
    import re
    import time
    # from datetime import date
    from selenium import webdriver
    import re
    import math
    from datetime import datetime
    from openpyxl import Workbook, load_workbook
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    import sqlitedbdatainsert
    # from dateutil import parser
except ImportError as e:
    # ImportErrorが発生した場合にモジュールをインストール
    if not moduleuse.install_missing_module(str(e)):
        print("An error occurred and the missing module could not be installed.")
        sys.exit(1)


def convert_currency_to_int(currency_str):
    # Use regular expression to remove non-digit characters (anything not a number or negative sign)
    cleaned_str = re.sub(r'[^\d-]', '', currency_str)
    
    # Convert the cleaned string to an integer
    return int(cleaned_str)

def model_validete_imput(text):
    models = [
        "デイトジャスト",
        "オイスター",
        "コスモグラフ",
        "シードゥエラー",
        "エクスプローラー",
        "GMTマスター",
        "GMTマスターII",
        "サブマリーナー",
        "ヨットマスター",
        "スカイドゥエラー",
        "エクスプローラーII",
        "エアキング",
    ]
    print("関数内のテキスト", text)
    pattern = r"\b\s+(\S+)\s+\b"
    # beltpattern = r"\[(.*?)\]|\((.*?)\)"
    beltpattern = r"\[(?:.*?)\]|\((?:.*?)\)"

    # ベルトと文字盤を抽出する。
    beltmatches = re.findall(beltpattern, text)
    if beltmatches:
        print("アイテム名が存在する")
        # beltmatches = re.sub(r'[\[\]()]', '', beltmatches[0])
    else:
        beltmatches = ["アイテム名を取得してください。"]
    # モデル名を抽出する。
    model = re.sub(beltpattern, "", text)
    print("モデル名", model)
    # [],()を正規表現で抽出する。

    items = {"model": model, "beltmatches": beltmatches}
    return items

def main():
    options = Options()
    options.add_argument("--auto-open-devtools-for-tabs")
    # 現在の日付を取得 
    today_date = datetime.now().date()
    # エクセルファイル作成
    file_name = f"EVANCEリスト{today_date}.xlsx"
    # datetimeオブジェクトに変換
    print(type(today_date))
    today_new_date_str = today_date.strftime('%Y/%m/%d')



    if not os.path.exists(file_name):
        # Excelブックの作成
        wb = Workbook()
        ws = wb.active
        # ヘッダー行を追加
        ws.append(
            [
             "モデル名",
             "リファレンス",
             "ブレスレット",
             "新品価格",
             "中古価格",
             "URL"
             ]
    )
        
    else:
        # ファイルが存在する場合は既存のファイルを読み込み
        wb = load_workbook(file_name)
        ws = wb.active


    num = 100
    # 数値の正規表現パターン
    number_pattern = re.compile(r"\d+")

    dinamicpagenum = 0
    # 合計数
    totalpagenum = 0
    # 最初に表示された件数を引いていって０ならば、処理を終える。
    # 最初の接続で表示された分だけループする。
    # https://kakaku.com/watch_accessory/watch/
    # 売れ筋ランキングURL
    # https://kakaku.com/watch_accessory/watch/itemlist.aspx?pdf_ma=5090
    url = f"https://kakaku.com/watch_accessory/watch/itemlist.aspx?pdf_ma=5090&pdf_so=h1"
    pal = "&pdf_pg="
    driver = webdriver.Chrome()
    driver.get(url)
    source_page_get =  BeautifulSoup(driver.page_source,"html.parser")
    # BeautifulSoupオブジェクトの作成

     # name属性が"frmComp"の要素を探す
    element = source_page_get.find(name='form', attrs={'name': 'frmComp'})
    print(element.text)
    
    # 全製品
    total_count = int(number_pattern.findall(element.find("p",class_="result").text.strip())[0])
    # 1ページ当たりの表示数
    print(total_count)
    onepage = 40
    # onepage = int(number_pattern.findall(element.find("span",class_="ec-font-bold").text.strip())[0])
    # 割り切れない場合、切り上げる。
    # pagenum = math.ceil((count/onepage)*10)/10
    # ループするページ数
    pagenum = math.ceil((total_count/onepage))
    print(f"トータル:{pagenum} 1ページ当たり{onepage}")
    # データべースのデータ数をトータル数で割る。
    dbcountitem = sqlitedbdatainsert.count_watch_items()
    print(dbcountitem)
    try:
        pagestartcount = math.ceil((total_count/dbcountitem))
    except ZeroDivisionError:
        pagestartcount = 0
    print(pagestartcount)
    # 一旦通信を終了
    driver.quit()
    # 以下取得したページ数だけ繰り返す。
    # 全ページのループ
    for pagestartcount in range(1,pagenum+1):
        # https://kakaku.com/watch_accessory/watch/itemlist.aspx?pdf_ma=5090&pdf_pg=1&pdf_pg=2
        url = f"https://kakaku.com/watch_accessory/watch/itemlist.aspx?pdf_ma=5090&pdf_so=h1"
        pal = "&pdf_pg="
        url = f"{url}{pal}{pagestartcount}"
        print(url)
        driver = webdriver.Chrome(options=options)
        driver.get(url)
        source_page_get = BeautifulSoup(driver.page_source,"html.parser")
        itemCatWrap = source_page_get.find_all("div",class_="itemCatWrap")
        # print(itemCatWrap)
       

       
        for item in itemCatWrap:
            a_tag_get = item.find("a")
            itemaccessurl = a_tag_get.get("href")
            print(f"タイムアウトになった場合→{url}")
            print(itemaccessurl)
            # タイムアウトになってもアクセスし続ける処理にする。
            driver.get(itemaccessurl)

            while True:
                # ページのロード状態をチェック
                if driver.execute_script('return document.readyState;') == 'complete':
                    break
                time.sleep(1)  # ロードが完了していなければ1秒待つ


            item_page_get = BeautifulSoup(driver.page_source,"html.parser")
            # 以下はitem_idにURLをいれている。
            item_id_match =  re.search(r'[A-Z]\d+', itemaccessurl)
            # print(item_id)
            if item_id_match:
                # ! DBに保存する。価格コムアイテムナンバー
                kakakukom_watch_id= item_id_match.group(0)
            else:
                kakakukom_watch_id = itemaccessurl
            titleBox =  item_page_get.find("div", id="titleBox")
            print(titleBox)
            titelname = item_page_get.find("div", id="titleBox").find("h2").get_text(strip=True)
            # 値段等
            try:
                productInfoBox = item_page_get.find("div", id="productInfoBox")
            # ref_pattern = r"\b(\d{4,6})([a-zA-Z]+)?\b"
                price_Item = productInfoBox.find("div",id="priceBox").find("div",class_="priceWrap").find("div",class_="subInfoObj1").find("p").find("span").get_text(strip=True)
                price_Item = convert_currency_to_int(price_Item)
                print(price_Item)
            except AttributeError:
                price_Item = ""
    

            try:
                use_price_Item = productInfoBox.find("p",class_="usedPrice").find("span",class_="usedPriceTxt").get_text(strip=True)
                use_price_Item = convert_currency_to_int(use_price_Item)
            except AttributeError:
                use_price_Item = ""
            print(use_price_Item)

            ref_pattern = r"\b(?:\d{4,6})(?:[a-zA-Z]+)?\b"
            refmatches = re.findall(
                ref_pattern,
                titelname,
                flags=re.UNICODE,
                )
            if refmatches:
                # !DBに保存する。リファレンスナンバー
                ref = refmatches[0]
            else:
                ref = ""
            # print(ref)
            # モデル名を取得するためにリファレンスナンバーを取り除く。
            model_matches = re.sub(ref_pattern,"",titelname)
            model_belt_bracelet_item = model_validete_imput(model_matches)        
            # !DBに保存する。モデル名　
            model_name = model_belt_bracelet_item["model"]
            # !DBに保存する。ダイアル名
            dial_name = model_belt_bracelet_item["beltmatches"][0]
            # !DBに保存する。アイテムURL
            set_url = itemaccessurl
            # !DBに保存する。ランキング
            ranking = item_page_get.find("a",class_="btn").find("span",class_="num").get_text(strip=True)   
            # ランキングを取得する。
            history_url = itemaccessurl+"pricehistory/"
            driver.get(history_url)
            history_page_soup = BeautifulSoup(driver.page_source,"html.parser")
            # 格コムのID：<class 'str'> モデル名:<class 'str'> リファレンス:<class 'str'> ブレスレット:<class 'dict'> ダイアル:<class 'list'> URL:<class 'str'>
            sqlitedbdatainsert.insert_watch_item(kakakukom_watch_id=kakakukom_watch_id,model_name=model_name,ref=ref,bracelet="",nowprice=price_Item, usenowprice=use_price_Item, dial=dial_name,url=set_url)
            
            while True:
                # ページのロード状態をチェック
                if driver.execute_script('return document.readyState;') == 'complete':
                    break
                time.sleep(1)  # ロードが完了していなければ1秒待つ

            try:
                # 価格推移グラフが存在しない
                history_table = history_page_soup.find("table",id="priceHistoryTbl1").find_all("tr")
                print(history_table)    
            except AttributeError:
                continue
           
            # print(history_table)
            # アイテムデータを入れる

            # 直近七日間ループする
            for datepriceitem in history_table:
                # CREATE TABLE IF NOT EXISTS weekly_reports (
                # report_id INTEGER PRIMARY KEY AUTOINCREMENT,
                # week_start_date TEXT NOT NULL,
                # ranking INTEGER,
                # summary TEXT,
                # price INTEGER,
                # kakakukom_watch_id TEXT NOT NULL,  
                # FOREIGN KEY (kakakukom_watch_id) REFERENCES watch_item(kakakukom_watch_id)
                try:
                    date = datepriceitem.find("td",class_="date").get_text(strip=True)
                    try:
                        # Trying to parse with non-zero-padded month and day
                        date_obj = datetime.strptime(date, '%Y年 %m月 %d日')
                    except ValueError:
                        # If fails, try to parse with a format that expects non-zero-padded month and day
                         date_obj = datetime.strptime(date, '%Y年 %m月%d日')
                    formatted_date = date_obj.strftime('%Y/%m/%d')
                    print(formatted_date)     
                    # date_obj = datetime.strptime(date, '%Y年 %m月 %d日')
                    # ranking = int(ranking)
                    price = datepriceitem.find("td",class_="price").get_text(strip=True)
                    # price = int(price)
                    price = convert_currency_to_int(price)
                    sqlitedbdatainsert.insert_weekly_report(week_start_date=formatted_date,ranking=ranking,summary=today_new_date_str,price=price,kakakukom_watch_id=kakakukom_watch_id)
                    # # def insert_weekly_report(week_start_date, ranking, summary, price, kakakukom_watch_id):
                except AttributeError as e:
                    print(f"Error processing data: {e}")  # Provides debug info about what went wrong
                    sqlitedbdatainsert.insert_weekly_report(week_start_date="",ranking="",summary=today_new_date_str,price="",kakakukom_watch_id=kakakukom_watch_id)
                    continue  # Skip this iteration and continue with the next one
                except ValueError as e:
                    print(f"Data conversion error: {e}")
                    sqlitedbdatainsert.insert_weekly_report(week_start_date="",ranking="",summary=today_new_date_str,price="",kakakukom_watch_id=kakakukom_watch_id)
                    continue

                #
                # print(date)
                print("---------------")
                


            





# mainメソッドを呼び出す
if __name__ == "__main__":
    main()