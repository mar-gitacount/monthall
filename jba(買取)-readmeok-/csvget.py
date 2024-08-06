from selenium import webdriver
import os
import json
from bs4 import BeautifulSoup
import re
import time
from datetime import datetime
from selenium import webdriver
import re
import math
from datetime import datetime
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import glob
import csv


def model_validete_imput(text):
    models = [
        "デイトジャスト",
        "オイスター",
        "コスモグラフ",
        "シードゥエラー",
        "エクスプローラー",
        "GMTマスター",
        "GMTマスターII",
        "サブマリーナー",
        "ヨットマスター",
        "スカイドゥエラー",
        "エクスプローラーII",
        "エアキング",
    ]
    print("関数内のテキスト", text)
    pattern = r"\b\s+(\S+)\s+\b"
    # beltpattern = r"\[(.*?)\]|\((.*?)\)"
    beltpattern = r"\[(?:.*?)\]|\((?:.*?)\)"

    # ベルトと文字盤を抽出する。
    beltmatches = re.findall(beltpattern, text)
    # モデル名を抽出する。
    model = re.sub(beltpattern, "", text)
    print("モデル名", model)
    # [],()を正規表現で抽出する。
    items = {"model": model, "beltmatches": beltmatches}
    return items

def main():
    # 現在の日付を取得 
    today_date = datetime.now().date()
    # 現在のディレクトリを格納する。
    path = os.getcwd()
    rowdata = []
    insertdata = []
    # CSVファイル作成
    # ファイルを取得し、利用おわったら済みフォルダに移動する。
    ref_pattern = r"\b(?:\d{4,7})(?:[a-zA-Z]+)?\s\b"
    gram_pattern = r"\b\d+.*?g\b"
    box_pattern = r"\b(BOX|ｹｰｽ)+\b"
    box_pattern = r"(BOX|ｹｰｽ|コマ)+"
    datepattern = r"\d+"

    # ref_pattern = r"\b(?:\d{4,6})(?:[a-zA-Z]+)*\s\b"

    # カレントディレクトリのCSVファイルを検索
    csv_files = glob.glob('csv/*.csv')

    excel_filename = ""

   # 検出されたCSVファイルのリストを表示
    for file_name in csv_files:
    #    print(file_name)
       use_csvfile = os.path.join(path,file_name)
           
    # 完了したcsvfile最後はdoneに移動する。
    excel_filename = re.findall(datepattern,use_csvfile)

    if not excel_filename:
        excel_filename = "errorname"
    else:
        excel_filename = "".join(excel_filename)
    excel_filename = "jba"+excel_filename + ".xlsx"
    header_items = ['会場', '開催日', '箱番号','番号','レーン名','品名','ブランド名','枠','脇石','重量','品番','','ランク','スタート価格','予想時刻','オークション結果','落札価格']
    if not os.path.exists(excel_filename):
        wb = Workbook()
        ws = wb.active
        # ヘッダーの数に合わせる
        
        ws.append(header_items)   
    else:
        # ファイルが存在する場合は既存のファイルを読み込み
        wb = load_workbook(file_name)
        ws = wb.active
    
    with open(use_csvfile, mode='r') as file:
        reader = csv.reader(file)
        header = next(reader)
        for row in reader:
            print(row)  
            filtered_items = [item for item in row if item not in [None, '', ' ']]
            rowdata.append(row)
        # print(rowdata)
        # itemsはループさせないで、順番で出力する
       
        
        for items in rowdata:
            itemdatas = [] * 20
            # 会場
            itemdatas.append(items[0])
            # 開催日
            itemdatas.append(items[1])
            # 箱番号
            itemdatas.append(items[2])
            # 番号
            itemdatas.append(items[3])
            # レーン名
            itemdatas.append(items[4])
            # 品名
            itemdatas.append(items[5])
            # ブランド名
            itemdatas.append(items[6])
            # 枠名
            itemdatas.append(items[7])
            # 脇石
            itemdatas.append(items[8])
            # 重量
            itemdatas.append(items[9])
            # 品番とその他を分ける
            refmatches = re.findall(ref_pattern,items[10],flags=re.UNICODE)
            if refmatches:
                parts = items[10].split() 
                ref = parts[0]
                another_datil = items[10].replace(ref,"")
                itemdatas.append(ref)
                itemdatas.append(another_datil)
            else:
                
                itemdatas.append("Noreference")
                itemdatas.append(items[10])
            # ランク
            itemdatas.append(items[11])
            # スタート価格
            itemdatas.append(items[12])
            # 予想時刻
            itemdatas.append(items[13])
            # オークション結果
            itemdatas.append(items[14])
            # 落札結果
            itemdatas.append(items[15])
            ws.append(itemdatas)
            continue




            # 列ごとのアイテムがある。
            # items = [item for item in items if item not in (None, '', ' ')]
            
            for index, item in enumerate(items):
                # n = 配列番号
                # itemdatas.append[items[n]　+ (ヘッダー数 - アイテム数)]
            #    array_index = index + (len(header_items)-len(items))
               
               array_index = index
               print(array_index)
               print(len(items),"これはアイテムの数")
               #以下には入らない。 
               if item is None or item == '' or item == ' ':
                  itemdatas.append(item)
                #   itemdatas.insert(array_index,array_index)
                  continue
                
               refmatches = re.findall(ref_pattern,item,flags=re.UNICODE)
               boxmatches = re.findall(box_pattern,item,flags=re.UNICODE)
               grammatches = re.findall(gram_pattern,item,flags=re.UNICODE)
               if boxmatches:
                   #いらないやつ
                   itemdatas = []
                   break

               if refmatches:
                  parts = item.split()
                  ref = parts[0]
                #   itemdatas.insert(array_index,ref)
                  itemdatas.append(ref)

                  # リファレンスナンバーを削除した他のデータ
                  another_datil = item.replace(ref,"")
                #   print(ref)
                # itemdatas.append("")
                  array_index += 1
                #   itemdatas.insert(array_index,another_datil)
                  itemdatas.append(another_datil)
               else:
                #  itemdatas.append("")
                #  itemdatas.insert(array_index,item)               
                 itemdatas.append(item)               
            
            ws.append(itemdatas)
            print(itemdatas)
            print("------------------")
            # print(ref)
            # print(itemdatas)
            
            # if itemdatas:
            #     insertdata.append()

        
        # ファイル名をコピーする。
        if not os.path.exists(use_csvfile):
            print(f"{use_csvfile}ある")
        else:
            print(f"{use_csvfile}はない。作る")
            # 元のファイルを取得する。
            make_csv_filename = os.path.basename(use_csvfile)
            # csvファイル作成する。
            with open(make_csv_filename,mode='w',newline='') as newfile:
                writer = csv.writer(newfile)
                writer.writerow(['会場', '開催日', '箱番号','番号','レーン名','品名','ブランド名','枠','脇石','重量','品番','','ランク','スタート価格','予想時刻','オークション結果','落札価格'])
                writer.writerows(insertdata)
    wb.save(excel_filename)
        




# mainメソッドを呼び出す
if __name__ == "__main__":
    main()