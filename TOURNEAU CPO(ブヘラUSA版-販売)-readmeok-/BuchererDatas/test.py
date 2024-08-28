
from whochedata_sqlite_data_insert import WhocheSqliteDataInsert
from sqlite_data_insert import SQLiteDataInsert
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import jpholiday
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import math
from openpyxl.styles import PatternFill
import re

def is_holiday_or_weekend(date):
    # 日付が日本の祝日かどうかを確認
    if jpholiday.is_holiday(date):
        return True
    # 日付が土曜日か日曜日かどうかを確認
    if date.weekday() >= 5:  # 5: Saturday, 6: Sunday
        return True
    return False

def save_logs_to_file(logs, file_path):
    # ここでアイテム一覧の配列を作ってしまう。
    # ここでの配列は二つで一つの二次元配列になる。

    with open(file_path, "a", encoding="utf-8") as file:
        file.write(str(logs) + "\n")
def modelnamereplace(value):
    return value

def materialnamereplace(value):
    return value


def number_to_column(number):
    """
    整数をExcelの列番号に変換する関数
    :param number: 整数
    :return: Excelの列番号（アルファベットの文字列）
    """
    column = ""
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        column = chr(65 + remainder) + column
    return column

def set_cell_color(sheet, row_number, column_letter, color):
    """
    指定された行と列のセルに色を設定する関数
    :param sheet: 対象のシート
    :param row_number: 行番号
    :param column_letter: 列のアルファベット
    :param color: 色（16進数カラーコード）
    """
    # 色を指定
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    print(row_number)
    print(column_letter)
    for col in range(1,column_letter+1):
        column = number_to_column(col)
        sheet[column + str(row_number)].fill = fill
    # column = ""

    # while column_letter > 0:
    #     column_letter, remainder = divmod(column_letter - 1, 26)
    #     column = chr(65 + remainder) + column
    
    # 指定された行と列のセルに色を設定
    sheet[column + str(row_number)].fill = fill

def main():
    
    # インスタンスを作成する。
    db_file = "TOURNEAU.db"
    # dbのファイルをわたす。
    whocequeryinstance = WhocheSqliteDataInsert(db_file)
    # テーブル一覧プロパティ
    tablename = whocequeryinstance.table_names
    today_date_and_time = datetime.now().strftime("%Y%m%d%H%M%S")
    items = whocequeryinstance.tablesdateill
    print(tablename,"テーブル名")
    # 現在の日付を取得
    today_date = datetime.now().strftime("%Y%m%d")
    file_name = f"BUCHRER CPOリスト{today_date}.xlsx"
    errot_file_name = f"HTMLエラーログ_{today_date_and_time}.txt"
  
    # print(items,"アイテム名")
    # テストで値を入れてみる
    # print(items['watch_item'].insert_data("watch_item",1, 2022, "Model 1", "REF123", "Bracelet 1", "Dial 1", "example.com"),"インスタンスたち")
    # items内にすべてのフィールドが格納されている。
    # 一致した場合、その日付のアイテムは除きたい。
    # 
    # jsondataloadmakeexcel.pyに移植する。
    week_items = []
    heder_rows =[]
    # 本日の日付を取得する
    today_date = datetime.now().date()
    # 月
    current_month = today_date.month
    # 変換後
    today_date_stringfomat = today_date.strftime('%Y/%m/%d')
    last_element = None
    modelnames = {"date":"DATE","datejust":"DJ","cosmograph-daytona":"DAYTONA","day-date":"DD","cellini":"その他","oyster-perpetual":"OP","yacht-master":"YACHT","sky-dweller":"SKY","sea-dweller":"SEA-DWELLER","turn-o-graph":"DJ","gmt-master-ii":"GMT","gmt-master":"GMT","submariner":"SUB","submariner-date":"SUB","deepsea":"DEEPSEA","explorer-ii":"EX","explorer":"EX"}
    material_codes = {"0":"SS","1":"SR","2":"SP","3":"SY","4":"SW","5":"RG","6":"PT","7":"TI","8":"YG","9":"WG"}
    # モデル名抽出。正規表現パターン
    modelpattern = r'https://www.bucherer.com/rolex-certified-pre-owned/watches/([^/]+)/'
    modelpattern = r'https://www\.tourneau\.com/watches/rolex-certified-pre-owned/([^/-]+)-*'
    numpattern = r'\d+'
    # 一回目の探索 items['weekly_reports'] weekdateフィールドを探索する　インスタンス　を使う。
    # trueなら配列にいれる？？
    # 本日の値が入っていた場合配列に日付を入れる

    week_date_count = 0
    while week_date_count == 0:
        week_date_count = items['weekly_reports'].datacountcheck(today_date_stringfomat,["weekdate"])
        # week_items.append(today_date_stringfomat) if week_date_count else print("アイテムなし")
        if week_date_count:
            week_items.append(today_date_stringfomat)
            save_logs_to_file(f"{today_date_stringfomat},はあります",errot_file_name)
        else:
            save_logs_to_file(f"{today_date_stringfomat},はありませんでした",errot_file_name)

        if week_date_count >= 1:
            break
        today_date = today_date-timedelta(1)
        today_date_stringfomat = today_date.strftime('%Y/%m/%d')
    
        # 本日のデータがない場合、today_dateに一番近い日付をいれる
       
    # week_items.append(today_date_stringfomat) if week_date_count else print("アイテムなし")


    
    # 以下月ごとに追加しているがこの処理を何とかしなければならない
    week_items.append("2024/07/26")
    week_items.append("2024/06/30")
    week_items.append("2024/05/23")
    week_items.append("2024/04/17")
    print(f"{week_date_count}は週があるかどうか")
    diff_date = today_date
    breakcount = 0
    while True:
        # 日付で検索、falseの場合、-1日
        # 検索でヒットした場合、その値の月を取得、そのまえの日と比べて月を跨いだら処理終了
        # マイナス7日でとる
        
        diff_date = diff_date-timedelta(1)

        if is_holiday_or_weekend(diff_date):
            print(f"{diff_date}は休日または週末です。")
            continue
        else:
            print(f"{diff_date}は平日です。")

        # 月を取得する  
        diff_date_month = diff_date.month
        # 文字列にする
        diff_date_string = diff_date.strftime('%Y/%m/%d')
        # 月が替わりかつそのアイテムがあればtrueになる。
        week_date_count = items['weekly_reports'].datacountcheck(diff_date_string ,["weekdate"])
        save_logs_to_file(f"{diff_date_string}は日付チェック",errot_file_name)
        # n日のデータが入り、かつ同じ月の場合処理を抜ける。
        if week_date_count and current_month == diff_date_month:
            continue
        #アイテムを加える。
        week_items.append(diff_date_string) if week_date_count else print("アイテムなし")
        breakcount += 1

        # if breakcount == 31:
        #     break

        if len(week_items) >= 4:
            break
    # 各週をチェックする。
    # save_logs_to_file(f"{week_items}に本日分が存在するか確認",errot_file_name)
    monthdata = whocequeryinstance.days_diffcheck('weekly_reports',"test",week_items)
    save_logs_to_file(f"{monthdata}はなにか",errot_file_name)
   
    reversed_week_items = week_items.reverse()
    print(week_items)
    new_index=[""]*len(week_items)
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(week_items)
    else:
        # ファイルが存在する場合は既存のファイルを読み込み
        wb = load_workbook(file_name)
        ws = wb.active

    # 月ごとファイル作成OK
    for id,dates in monthdata.items():
         print(f"呼び出し元→ID: {id}, Dates: {dates}")
         for date in dates:
            index = week_items.index(date)
            print(f"{date}日付 番号:{id}")
            save_logs_to_file(f"呼び出し元→ID: {id}, Dates: {dates}",errot_file_name)
            print("-------------")
            new_index.insert(index,id)
        # データを加える
         print(new_index)
         ws.append(new_index)
         new_index=[""]*len(week_items)

    # 一旦エクセル保存する。
    wb.save(file_name)
    # 今週、先週差分チェック
    # 最初と最後のアイテムを抽出する。
    if len(week_items)>=2:
        last_element = week_items[-1]
        second_last_element = week_items[-2]
        week_items = week_items[-2:]
        save_logs_to_file(f"先週？{last_element}先々週?{second_last_element}",errot_file_name)
    today_date_stringfomat = today_date.strftime("%Y-%m-%d")


    new_sheet = wb.create_sheet(title=today_date_stringfomat)
    # 入りと売れシート
    week_saleandbuy_sheet = wb.create_sheet(title=today_date_stringfomat+"出入り")
    sale_items =[]
    buy_items =[]
    monthdata = whocequeryinstance.days_diffcheck('weekly_reports',"test",week_items)
    whocequeryinstance.close_connection()

    

    
    # 再度インスタンスを作成する
    whocequeryinstance_watch_item_use = WhocheSqliteDataInsert(db_file)
    # テーブル一覧プロパティ
    watch_item_use_tablename = whocequeryinstance_watch_item_use.table_names
    watch_item_use_items = whocequeryinstance_watch_item_use.tablesdateill
    # 時計一覧のインスタンスを作成する。
    bucherer_table_instance = watch_item_use_items["watch_item"]
    # 以下から↓
    new_sheet.append(["BUCHERER CPO LIST"])
    # 以下円価格はwebdriverでしらべる
    url = f"https://www.google.com/search?q=%E3%82%B9%E3%82%A4%E3%82%B9%E3%83%95%E3%83%A9%E3%83%B3&oq=%E3%82%B9%E3%82%A4%E3%82%B9%E3%83%95%E3%83%A9%E3%83%B3&gs_lcrp=EgZjaHJvbWUqDggAEEUYJxg7GIAEGIoFMg4IABBFGCcYOxiABBiKBTIHCAEQABiABDIHCAIQABiABDIHCAMQABiABDIHCAQQABiABDIHCAUQABiABDIHCAYQABiABDIHCAcQABiABDIHCAgQABiPAjIHCAkQABiPAtIBCDIwMDZqMGo3qAIAsAIA&sourceid=chrome&ie=UTF-8"

    options = Options()
    options.add_argument("--headless")  # ヘッドレスモードを有効にする
    # SeleniumのWebDriverを初期化
    driver = webdriver.Chrome(options=options)  # または他のブラウザに合わせて選択
     # URLを開く
    chf_rate = 0

    try:
        # Googleでスイスフランから日本円の為替レートを検索
       driver.get('https://www.google.com')
       search_box = driver.find_element(By.NAME, 'q')
       search_box.send_keys('USD to JPY')
       search_box.submit()
    
       # 結果が表示されるまで待機
       time.sleep(3)
       page_source = driver.page_source
       soup= BeautifulSoup(page_source, "html.parser")
    #    gettag = soup.find_all("div",class_="dDoNo")
    #    gettag = soup.find_all("span",class_="DFlfde")
       gettag = soup.find("span",class_="SwHCTb")
       print(f'The current exchange rate for CHF to JPY is: {gettag}')
       chf_rate_number = float(gettag.get_text(strip=True))
       chf_rate = math.floor(chf_rate_number)
       save_logs_to_file(chf_rate,errot_file_name)
    except Exception as e:
         # エラーが発生した場合はログに記録
         error_message = f"エラーが発生しました: {str(e)}"
         print(error_message)

         save_logs_to_file(error_message, errot_file_name)
    finally:
        # WebDriverを終了
        save_logs_to_file("レート取得完了",errot_file_name)
        driver.quit()
    

    
    new_sheet.append(["","商品番号","モデル","年式","サイズ","素材","Ref.","ブレスレット","ダイアル","再販価格",f"円価格(CHF＝{chf_rate}円)","前価格","差額"])
    # 一週前のデータと比較する処理。
    indexnum = 1
    # new_index=[""]*15
    for id,dates in monthdata.items():
        print(f"呼び出し元→ID: {id}, Dates: {dates}")
        # 呼び出し元→ID: 1411-153-2, Dates: ['2024/05/27', '2024/05/30']
        # 日付入稿ループ
        # The loop is in ascending order.
        new_index=[""]*15
        save_logs_to_file(f"呼び出し元→ID: {id}, Dates: {dates}",errot_file_name)
        #  week_date_count = items['weekly_reports'].datacountcheck(diff_date_string ,["weekdate"])
        # 前の週の値段と比較する
        if len(dates) == 2:
            # 
            before_date_price = items['weekly_reports'].serachitem([id,dates[0]],["bucherer_watch_id","weekdate"])
            before_date_price = list(before_date_price[0])
            after_date_price = items['weekly_reports'].serachitem([id,dates[1]],["bucherer_watch_id","weekdate"])
            after_date_price = list(after_date_price[0])
            save_logs_to_file(f"{before_date_price}が{dates[0]}の金額データ",errot_file_name)
            save_logs_to_file(f"{after_date_price}が{dates[1]}の金額データ",errot_file_name)
            itemvalue = bucherer_table_instance.serachitem([id],["bucherer_watch_id"])
            save_logs_to_file(itemvalue,errot_file_name)
            
            itemvalue = list(itemvalue[0])

            # モデル名取得。
            modelmatch = re.search(modelpattern,itemvalue[8])
            if modelmatch:
                modelValue  = modelmatch.group(1)
            else:
                modelValue = itemvalue[8]
            # 辞書チェック
            if modelValue in modelnames:
                model_input_value = modelnames[modelValue]
            else:
                model_input_value = "その他"
            
            material_code_match = re.search(numpattern,itemvalue[5])
            if material_code_match:

                material_code = material_code_match.group()
                print(material_code)
                lastnum = material_code[-1:]
                if lastnum in material_codes:
                    material_code = material_codes[lastnum]
                
            else:
                material_code = "エラー素材コードを手で入力してください。"

            
            new_index.insert(0,indexnum)
            # アイテムナンバー
            new_index.insert(1,itemvalue[1])
            # モデル名
            # new_index.insert(2,itemvalue[3])
            new_index.insert(2,model_input_value)
            # 年代
            new_index.insert(3,itemvalue[2])
            # サイズ
            new_index.insert(4,itemvalue[4])
           
            # 素材あとで数字に帰る
            # new_index.insert(5,itemvalue[4])
            # 素材コード
            new_index.insert(5,material_code)
            new_index.insert(6,itemvalue[5])
            # ブレスレット
            new_index.insert(7,itemvalue[6])
            # ダイアル
            new_index.insert(8,itemvalue[7])
            # 値段
            # new_index.insert(9,before_date_price[3])
           
            # 値段
            new_index.insert(9,after_date_price[3])
            # 日本円にする
            enchange = after_date_price[3]*chf_rate_number
             # 差額の価格
            last_row = new_sheet.max_row
            # new_index.insert(10,f"=J{last_row+1}*{chf_rate}")
            new_index.insert(10, f'=TEXT(J{last_row + 1}, "¥#,##0") * {chf_rate}')

            new_index.insert(11,before_date_price[3])
           
            print(last_row)
            new_index.insert(12,f"=J{last_row+1}-L{last_row+1}")
            
            # url
            new_index.insert(13,itemvalue[8])
            # URL
            # save_logs_to_file(itemvalue,errot_file_name)(10,itemvalue[8])
            new_sheet.append(new_index)
            # new_index=[""]*len(15)
            new_index=[""]*15
            indexnum +=1 
        # ここでまた最新の日を判定しなければならない。
        # データが一つの時。
        else:
            # 新しく入荷したもの
            # last_elementは二個以上のときに実行されるので、以下エラーになる。
            # 比べる月の合計数=２以上で、ループ内の比較アイテムが一個になる場合は動く
            if last_element == dates[0]:
                # 決算シートに入荷として代入する。
                before_date_price = items['weekly_reports'].serachitem([id,dates[0]],["bucherer_watch_id","weekdate"])
                before_date_price = list(before_date_price[0])
                # 時計一覧テーブルを検索している。
                itemvalue = bucherer_table_instance.serachitem([id],["bucherer_watch_id"])
                itemvalue = list(itemvalue[0])
                
                save_logs_to_file(itemvalue,errot_file_name)
                #   0        1           2                    3                     4       5     6           7                         8                        
                # [929, '1411-689-9', '2016', 'Datejust  Certified Pre-Owned', '178383', '31mm', '3', 'sales box (original)', 'https://www.bucherer.com/rolex-certified-pre-owned/watches/datejust/1411-689-9.html']
                # アイテムナンバー
                modelmatch = re.search(modelpattern,itemvalue[8])
                if modelmatch:
                    modelValue  = modelmatch.group(1)
                else:
                    modelValue = itemvalue[8]
                # 辞書チェック
                if modelValue in modelnames:
                    model_input_value = modelnames[modelValue]
                else:
                    model_input_value = "その他"
                print(itemvalue,"全データ")
                print(itemvalue[5],"これは素材コードの元")
                material_code_match = re.search(numpattern,itemvalue[5])
                if material_code_match:
                    
                    material_code = material_code_match.group()
                    lastnum = material_code[-1:]
                    if lastnum in material_codes:
                        material_code = material_codes[lastnum]
                else:
                    material_code = "エラー素材コードを手で入力してください。"
                new_index.insert(0,indexnum)
                 # アイテムナンバー
                new_index.insert(1,itemvalue[1])
                 # モデル名
                new_index.insert(2,model_input_value)
                # 年代
                new_index.insert(3,itemvalue[2])
                # サイズ
                new_index.insert(4,itemvalue[4])
                # 素材あとで数字に変える
                new_index.insert(5,material_code)
                new_index.insert(6,itemvalue[5])
                # ブレスレット
                new_index.insert(7,itemvalue[6])
                # ダイアル
                new_index.insert(8,itemvalue[7])
                # 値段
                new_index.insert(9,before_date_price[3])
                last_row = new_sheet.max_row
                enchange = before_date_price[3]*chf_rate_number
                # レート計算
                # new_index.insert(10,f"=J{last_row+1}*{chf_rate}")
                # 出入りシートに入れるための値
                sale_items.append(new_index)
                new_index.insert(10, f'=TEXT(J{last_row + 1}, "¥#,##0") * {chf_rate}')

                new_index.insert(11,0)
                 # url
                new_index.insert(13,itemvalue[8])
                save_logs_to_file(f"{before_date_price}が{dates[0]}の金額データ",errot_file_name)
                new_sheet.append(new_index)
                # new_index=[""]*len(15)
                new_index=[""]*15

                indexnum +=1 
            # falseの場合は、売れなので、それを代入
            # とりあえず配列に退避させる
            else:
                # 決算シートに入荷として代入する。
                before_date_price = items['weekly_reports'].serachitem([id,dates[0]],["bucherer_watch_id","weekdate"])
                before_date_price = list(before_date_price[0])
                # 時計一覧テーブルを検索している。
                itemvalue = bucherer_table_instance.serachitem([id],["bucherer_watch_id"])
                itemvalue = list(itemvalue[0])
                save_logs_to_file(itemvalue,errot_file_name)
                modelmatch = re.search(modelpattern,itemvalue[8])
                if modelmatch:
                    modelValue  = modelmatch.group(1)
                else:
                    modelValue = itemvalue[8]
                # 辞書チェック
                if modelValue in modelnames:
                    model_input_value = modelnames[modelValue]
                else:
                    model_input_value = "その他"

                print(itemvalue,"全データ")
                print(itemvalue[4],"これは素材コードの元")
                material_code_match = re.search(numpattern,itemvalue[5])
                if material_code_match:
                    material_code = material_code_match.group()
                    lastnum = material_code[-1:]
                    if lastnum in material_codes:
                        material_code = material_codes[lastnum]
                else:
                    material_code = "エラー素材コードを手で入力してください。"
                #   0        1           2                    3                     4       5     6           7                         8                        
                # [929, '1411-689-9', '2016', 'Datejust  Certified Pre-Owned', '178383', '31mm', '3', 'sales box (original)', 'https://www.bucherer.com/rolex-certified-pre-owned/watches/datejust/1411-689-9.html']
                # アイテムナンバー
                new_index.insert(0,indexnum)
                # アイテムナンバー
                new_index.insert(1,itemvalue[1])
                # モデル名
                new_index.insert(2,model_input_value)
                # 年代
                new_index.insert(3,itemvalue[2])
                # サイズ
                new_index.insert(4,itemvalue[4])
                # 素材あとで数字に帰る
                new_index.insert(5,material_code)
                new_index.insert(6,itemvalue[5])
                # ブレスレット
                new_index.insert(7,itemvalue[6])
                # ダイアル
                new_index.insert(8,itemvalue[7])
                # 値段
                new_index.insert(9,before_date_price[3])
                last_row = new_sheet.max_row
                enchange = before_date_price[3]*chf_rate_number
                # レート計算
                # new_index.insert(10,f"=J{last_row+1}*{chf_rate}")
                # 出入りシートに入れるための値
                buy_items.append(new_index)
                new_index.insert(10, f'=TEXT(J{last_row + 1}, "¥#,##0") * {chf_rate}')

                new_index.insert(11,0)
                # url
                new_index.insert(13,itemvalue[8])
                save_logs_to_file(f"{before_date_price}が{dates[0]}の金額データ",errot_file_name)
                # new_sheet.append(new_index)
                # new_index=[""]*len(15)
                new_index=[""]*15

                # indexnum +=1 

    
    

    bucherer_table_instance.close_connection()
    # 出入りシート作成
    week_saleandbuy_sheet.append(["","商品番号","モデル","年式","サイズ","素材","Ref","ブレスレット","ダイアル","再販価格",f"円価格(CHF＝{chf_rate}円)"])
    
    for sale_item in sale_items:
        week_saleandbuy_sheet_last_row = week_saleandbuy_sheet.max_row
        week_saleandbuy_sheet_insertvalue=[""]*11
        color = "D9D9D9"
        # 下に行がたされてしまうので、あとで考える
        # set_cell_color(week_saleandbuy_sheet,week_saleandbuy_sheet_last_row+1,1,color)
        week_saleandbuy_sheet_insertvalue.insert(1,sale_item[1])
        modelreplace = modelnamereplace(sale_item[2])
        # モデル名
        week_saleandbuy_sheet_insertvalue.insert(2,modelreplace)
        # 年式
        week_saleandbuy_sheet_insertvalue.insert(3,sale_item[3])
        # サイズ
        week_saleandbuy_sheet_insertvalue.insert(4,sale_item[4])
        # 素材
        material = materialnamereplace(sale_item[5])
        week_saleandbuy_sheet_insertvalue.insert(5,material)
        # リファレンス
        week_saleandbuy_sheet_insertvalue.insert(6,sale_item[5])
        
        # ブレスレット
        week_saleandbuy_sheet_insertvalue.insert(7,sale_item[7])
        # ダイアル
        week_saleandbuy_sheet_insertvalue.insert(8,sale_item[8])
       
        # 再販化価格(CHF)
        week_saleandbuy_sheet_insertvalue.insert(9,sale_item[9])
        # new_index.insert(10, f'=TEXT(J{last_row + 1}, "¥#,##0") * {chf_rate}')
        week_saleandbuy_sheet_insertvalue.insert(10,f'=J{week_saleandbuy_sheet_last_row + 1}*{chf_rate}')
        week_saleandbuy_sheet.append(week_saleandbuy_sheet_insertvalue)
    week_saleandbuy_sheet_insertvalue=[""]*11
    week_saleandbuy_sheet_last_row = week_saleandbuy_sheet.max_row
    week_saleandbuy_sheet_insertvalue.insert(9,f'=SUM(J2:J{week_saleandbuy_sheet_last_row})')
    week_saleandbuy_sheet_insertvalue.insert(10,f'=SUM(K2:K{week_saleandbuy_sheet_last_row})')
    week_saleandbuy_sheet.append(week_saleandbuy_sheet_insertvalue)
    
    save_logs_to_file(f"売れた数→{len(buy_items)} 買った数→{len(sale_items)}",errot_file_name)
    # 最初の行を更新する
    week_saleandbuy_sheet_min_row = week_saleandbuy_sheet.max_row+1
    for buy_item in buy_items:
        week_saleandbuy_sheet_last_row = week_saleandbuy_sheet.max_row
        week_saleandbuy_sheet_insertvalue=[""]*11
        color = "D9D9D9"
        # 下に行がたされてしまうので、あとで考える
        # set_cell_color(week_saleandbuy_sheet,week_saleandbuy_sheet_last_row+1,1,color)
        week_saleandbuy_sheet_insertvalue.insert(1,buy_item[1])
        modelreplace = modelnamereplace(buy_item[2])
        # モデル名
        week_saleandbuy_sheet_insertvalue.insert(2,modelreplace)
        # 年式
        week_saleandbuy_sheet_insertvalue.insert(3,buy_item[3])
        # サイズ
        week_saleandbuy_sheet_insertvalue.insert(4,buy_item[4])
        # 素材
        material = materialnamereplace(buy_item[5])
        week_saleandbuy_sheet_insertvalue.insert(5,material)
        # リファレンス
        week_saleandbuy_sheet_insertvalue.insert(6,buy_item[5])
        
        # ブレスレット
        week_saleandbuy_sheet_insertvalue.insert(7,buy_item[7])
        # ダイアル
        week_saleandbuy_sheet_insertvalue.insert(8,buy_item[8])
       
        # 再販化価格(CHF)
        week_saleandbuy_sheet_insertvalue.insert(9,buy_item[9])
        # new_index.insert(10, f'=TEXT(J{last_row + 1}, "¥#,##0") * {chf_rate}')
        week_saleandbuy_sheet_insertvalue.insert(10,f'=J{week_saleandbuy_sheet_last_row + 1}*{chf_rate}')
        week_saleandbuy_sheet.append(week_saleandbuy_sheet_insertvalue)


    
    week_saleandbuy_sheet_insertvalue=[""]*11
    week_saleandbuy_sheet_insertvalue.insert(9,f'=SUM(J{week_saleandbuy_sheet_min_row}:J{week_saleandbuy_sheet_last_row})')
    week_saleandbuy_sheet_insertvalue.insert(10,f'=SUM(K{week_saleandbuy_sheet_min_row}:K{week_saleandbuy_sheet_last_row})')
    week_saleandbuy_sheet.append(week_saleandbuy_sheet_insertvalue)
    



    # 合計金額を入れ込む


    # # 行方向
    # for row_index ,row_data in enumerate(new_sheet,start=1):
    #     if row_index <3 :
    #         continue
    #     value = row_data[9]
    #     cell = ws.cell(row=row_index,column=10,value=value)
    #     formula = f'=TEXT({cell.coordinate}, "¥#,##0")'
    #     cell.formula = formula


    wb.save(file_name)
    # for item in week_items:
    #     print(f"{week_items}は週ぜんぶ　")
    #     whocequeryinstance.days_diffcheck('weekly_reports',item,week_items)
    #     print(f"{item}は週ごと")
    #     # ここから各週ごとのアイテムをチェックする
    #     # アイテム(n)→週分検索する。

    # 
    for i in range(4):
        print(i)
        print(today_date-timedelta(1))
        today_date = today_date-timedelta(1)
        today_date_stringfomat = today_date.strftime('%Y/%m/%d')
        print(today_date_stringfomat)

    
    dbinsert_datagetnow = datetime.strptime(today_date,'%Y/%m/%d').date()
    print(dbinsert_datagetnow)
    # for i in range(4):
    #     print(i+1)
    #     adddate = i+1
    #     # 七日マイナスする。
    #     # 計算式= n日 - (7*(4-k))
    #     print(adddate)
    #     print(type(dbinsert_datagetnow))
                            
    #     print(type(dbinsert_datagetnow))
    #     print(dbinsert_datagetnow)
    #     dbinsert_datagetnow_minusdate = dbinsert_datagetnow - timedelta(days=7*adddate)
    #     # 月が違う場合処理してbreakする。
    #     if dbinsert_datagetnow.month == dbinsert_datagetnow_minusdate.month:
    #          print("同じ月")
    #     else:
    #          print("違う月、ここでおわり")
    #          print(f"{dbinsert_datagetnow.month}と{dbinsert_datagetnow_minusdate.month}")
    #          # dbinsert_datagetnow_minusdate = dbinsert_datagetnow - timedelta(days=1)
    #          print(dbinsert_datagetnow_minusdate)
                            

    fieldcount = items['watch_item'].fieldcountAllcountcheck()
   
    
    
    
    fieldcount = items['watch_item'].fieldcountAllcountcheck()
    value = [f"test" ,2022, "Model 2", "REF123", "Bracelet 1", "Dial 1", "example.com"]
    jikou = items['watch_item'].insert_data(value)
    weeklydatas = ["2024/05/23","0","4000","1348-535-3"]
    weekinserttest = items['weekly_reports'].insert_data(weeklydatas)


    print(items,"テーブル一覧")
    print("→は売れた数",buy_items)
    print("→は入荷した数",sale_items)
    # テーブル数を確認して自動増分する。
    
 
    # jikou = whocequeryinstance.insertsuper(value)

    # jikou.close_connection()


if __name__ == "__main__":
    main()