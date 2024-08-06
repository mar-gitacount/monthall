import time
from datetime import datetime,timedelta
import sqlitedbdatainsert
from openpyxl import Workbook, load_workbook
import os
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference
def number_to_column(number):
    """
    整数をExcelの列番号に変換する関数
    :param number: 整数
    :return: Excelの列番号（アルファベットの文字列）
    """
    column = ""
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        column = chr(65 + remainder) + column
    return column



def set_cell_color(sheet, row_number, column_letter, color):
    """
    指定された行と列のセルに色を設定する関数
    :param sheet: 対象のシート
    :param row_number: 行番号
    :param column_letter: 列のアルファベット
    :param color: 色（16進数カラーコード）
    """
    # 色を指定
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    print(row_number)
    print(column_letter)
    for col in range(1,column_letter+1):
        column = number_to_column(col)
        sheet[column + str(row_number)].fill = fill
    # column = ""

    # while column_letter > 0:
    #     column_letter, remainder = divmod(column_letter - 1, 26)
    #     column = chr(65 + remainder) + column
    
    # 指定された行と列のセルに色を設定
    sheet[column + str(row_number)].fill = fill



def get_max_column(sheet, row_number):
    """
    指定された行の最大の列を取得する関数
    :param sheet: 対象のシート
    :param row_number: 最大の列を取得する行番号
    :return: 最大の列番号
    """
    max_column = 0

    # 指定された行の最大の列番号を取得
    for cell in sheet[row_number]:
        if cell.value is not None:
            max_column = cell.column

    return max_column



def workbookmake(bookname):
    if not os.path.exists(bookname):
        wb = Workbook()
        ws = wb.active
        return wb
    else:
        wb = Workbook()
        ws = wb.active
        return wb
def sheet_exists(wb, sheet_name):
    return sheet_name in wb.sheetnames


def main():
    today_date = datetime.now().date()
    file_name = f"価格コムリスト{today_date}.xlsx"
    datediff = 0
    # 日付がみつかるまで検索
    while True:
        # Print the type of today_date
        print(type(today_date))
        today_date = today_date + timedelta(days=datediff)
        # Convert datetime object to string in the format 'YYYY/MM/DD'
        today_new_date_str = today_date.strftime('%Y/%m/%d')
        print(today_new_date_str)
        # 日付データを確認して0以上の場合、処理を抜ける。
        queryitems = sqlitedbdatainsert.datagetcount(today_new_date_str)
        
        # 本日の日付で検索する。なければ、マイナス1日ずつして検索する。
        # データの長さを判定する。

        if not queryitems:
            # 日付変数を減らす
            datediff -= 1
            continue
        else:
            file_name = f"価格コム売上ランキングリスト{today_date}.xlsx"
            wb = workbookmake(file_name)
            break
    # print(queryitems)
    # デイリーでとるデータは、-1日＆-7日
    dayly_data_date = today_date+timedelta(days=-1)

    dayly_beforeseven_date = dayly_data_date+timedelta(days=-7)

    print(dayly_data_date)
    print(dayly_beforeseven_date)
    # タプル内の文字列を取り出してリストに格納する
    print(f"{queryitems}はランキング配列")
    # strings = [t[0] for t in queryitems]
    # queryitemsはリスト
    # n日分の検索データが抽出される。
    # 日付でソートされたデータ群で更に検索する。
    for item in queryitems:
        # item = kakakukom_watch_id になっているので、時計テーブルもこれで検索する
        itemstatus = sqlitedbdatainsert.watch_item_only_get(item)
        itemstatus = list(itemstatus[0])
        
        # fixed_tuple = [item if item else None for item in itemstatus[0]]
        
        # itemstatus = [t or t in itemstatus]
        # アイテムごとの値が格納されている。シリアルナンバーのみが格納されている。
        weekly_items = sqlitedbdatainsert.weekly_reports_only_get(item,today_new_date_str)
        # シートを作成してかえってきたデータを取得する。
        print(item)
        print(itemstatus)
        # 色を設定する行番号と列のアルファベットを指定
        row_number = 1
        column_letter = 'A'
        color = "FFFF00"  # 赤色の例
        # ワークシートが存在する
        if sheet_exists(wb,item):
            ws = wb.active
            ws.title = item
        else:
        # ワークシート作成する
            ws = wb.create_sheet(title=itemstatus[3])
        ws.append(
            [
             itemstatus[2],
             itemstatus[3],
             itemstatus[5],
             itemstatus[6]
            ]
            )
        set_cell_color(ws, ws.max_row, get_max_column(ws,ws.max_row), color)
        # ここに最終行をいれる。
        print(f"何行目→？{ws.max_row}最大列{get_max_column(ws,ws.max_row)}")
        ws.append(
        [
             "日付",
             "値段",
             "差額"
             ]
             )
        color = "D9D9D9"
        set_cell_color(ws, ws.max_row, get_max_column(ws,ws.max_row), color)
        print(f"次は何行目→{ws.max_row} 最大列{get_max_column(ws,ws.max_row)}")
        diff_input_cell = 3
        min_low = ws.max_row
        min_low += 1
        max_low = ws.max_row
        max_low += 1
        
        min_column = 1
        max_column = 2

        for weekly_item in weekly_items:
            day_keys = list(weekly_item.keys())[0]
            price_into = weekly_item[day_keys]
            price = price_into['price']
            
            #次のセル 
            next_cell = diff_input_cell + 1
            ws.append([day_keys,price,f"=B{diff_input_cell}-B{next_cell}"])
            print(day_keys)
            print(price)
            print("------------")
            # 行を追加
            max_low += 1
            diff_input_cell += 1


        # 折れ線グラフを作成
        chart = LineChart()
        chart.title = "最安値"
        chart.style = 10
        chart.y_axis.title = '値段'
        chart.x_axis.title = '日付'
 

        # 値段データ
        data = Reference(ws,min_col=2,min_row=min_low,max_col=2,max_row=max_low)
        # 日付データ
        categories = Reference(ws, min_col=1, min_row=min_low, max_row=max_low)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # グラフをシートに追加
        ws.add_chart(chart, "E5")

        
        # ウィークリーテーブルを検索する。
        # itemと日付を渡す。そうすると、当日の値段が返ってくる。それが完了したら、-七日を検索してみる。エラーになる可能性はある。
    wb.save(file_name)   

    # エクセルに入稿する。
    # 日付クエリ実行する。
    # クエリをループして、そのなかで週ごとに取得する。




   




    # mainメソッドを呼び出す
if __name__ == "__main__":
    main()