from selenium import webdriver
import os
import json
from bs4 import BeautifulSoup
import re
import time
from datetime import datetime
from selenium import webdriver
import re
import math
from datetime import datetime
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
def main():
    # 現在の日付を取得 
    today_date = datetime.now().date()
    # エクセルファイル作成
    file_name = f"EVANCEリスト{today_date}.xlsx"

    if not os.path.exists(file_name):
        # Excelブックの作成
        wb = Workbook()
        ws = wb.active
        # ヘッダー行を追加
        ws.append(
            [
             "モデル名",
             "リファレンス",
             "ブレスレット",
             "新品価格",
             "中古価格",
             "URL"
             ]
    )
        
    else:
        # ファイルが存在する場合は既存のファイルを読み込み
        wb = load_workbook(file_name)
        ws = wb.active


    num = 100
    # 数値の正規表現パターン
    number_pattern = re.compile(r"\d+")

    dinamicpagenum = 0
    # 合計数
    totalpagenum = 0
    # 最初に表示された件数を引いていって０ならば、処理を終える。
    # 最初の接続で表示された分だけループする。
    url = f"https://evance.co.jp/products/list?category_id=107&stock%5B0%5D=1&pageno=1"
    driver = webdriver.Chrome()
    driver.get(url)
    source_page_get =  BeautifulSoup(driver.page_source,"html.parser")
    totalget = source_page_get.find("section",class_="area_paging")
    
    
    count = int(number_pattern.findall(totalget.find("p",class_="count").text.strip())[0])
    onepage = int(number_pattern.findall(totalget.find("span",class_="ec-font-bold").text.strip())[0])
    
    print(f"トータル:{count} 1ページ当たり{onepage}")
    # pagenum = math.ceil((count/onepage)*10)/10
    # ループするページ数
    pagenum = math.ceil((count/onepage))
    # 一旦通信を終了
    driver.quit()
    # 以下取得したページ数だけ繰り返す。
    for i in range(1,pagenum+1):
        url = f"https://evance.co.jp/products/list?category_id=107&stock%5B0%5D=1&pageno={i}"
        driver = webdriver.Chrome()
        driver.get(url)
        source_page_get = BeautifulSoup(driver.page_source,"html.parser")
        
        list_item_main = source_page_get.find("ul",class_="list_item_main")
        item_atags = list_item_main.find_all("a")

        for a in item_atags:
            print(a)
            try:
                itemaccessurl = a.get("href")
                driver.get(itemaccessurl)
                wait = WebDriverWait(driver, 10)  # 最大10秒待つ
                wait.until(EC.presence_of_element_located((By.TAG_NAME, 'html')))
                itempage = BeautifulSoup(driver.page_source,"html.parser")

                area_register = itempage.find("div",class_="area_register")
                
                infodateil = itempage.find("div",class_="area_information_detail")
                # print(infodateil)
            except Exception as e:
                continue
            #  onepage = int(number_pattern.findall(totalget.find("span",class_="ec-font-bold").text.strip())[0])
            
            model_tag = itempage.find("p",class_="ttl",text="モデル")
            # もしかしたら、map処理できるかも！！
            # ?モデル名
            try:
                model_name = model_tag.find_next_sibling().text.strip()
            except Exception as e:
                model_name = ""
            
            ref_tag = itempage.find("p",class_="ttl",text="型番")
            print(model_name)
            print(ref_tag)
            
            # ?リファレンスナンバー
            ref_no = ref_tag.find_next_sibling().text.strip()
            
            # ?値段
            price = itempage.find("p",class_="price").text.strip()
            
            # ?新品判定
            new_status = area_register.find("p",class_="status_new")

            # ?未使用品判定
            unused_status = area_register.find("p",class_="status_unused")

            # ?中古判定
            used_status = area_register.find("p",class_="status_used")
        
            # ?値段を取得。
            price = re.search(r"￥([\d,]+)", price).group(1) 
            
            # ?ブレスレット
            bracelet_tag = itempage.find("p",class_="ttl",text="ベルトタイプ")
            if bracelet_tag != None:
                bracelet = bracelet_tag.find_next_sibling().text.strip()
            else:
                bracelet = ""

            print(model_name)
            print(ref_no)           
            print(price)
            print(new_status)
            print(used_status)
            print(unused_status)

            # データ入力する
            if new_status != None or unused_status != None:
            #           [
            #  "モデル名",
            #  "リファレンス",
            #  "ブレスレット",
            #  "新品価格",
            #  "中古価格",
            #  "URL"
            #  ]
                
                datarow = [model_name,ref_no,bracelet,price,"",itemaccessurl]
            else:
                datarow = [model_name,ref_no,bracelet,"",price,itemaccessurl]
            ws.append(datarow)
            print("-------------------------------")
        wb.save(file_name)
        




# mainメソッドを呼び出す
if __name__ == "__main__":
    main()