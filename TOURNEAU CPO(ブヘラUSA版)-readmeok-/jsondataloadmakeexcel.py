import os
import json
import csv
from bs4 import BeautifulSoup
import requests
import re
import sys
from openpyxl import Workbook, load_workbook
from datetime import datetime
import time
import os
import json
from datetime import datetime
# JSONファイルのパス
BuchererMainDatasjson = "Buchererjson/BuchererMainDatas.json"
# JSONデータの読み込み
with open(BuchererMainDatasjson, "r", encoding="utf-8") as file:
    main_json_data = json.load(file)



def extract_single_digit(text):
    for char in text:
        if char.isdigit():
            return char
    return None

# リファレンスでチェックする。
def get_item_details(json_data, ref):
    item_details = []
    master_data = json_data.get("master", {})
    for item_key, item_data in master_data.items():
        if item_data.get("ref") == ref:
            item_details.append(item_data)
    return item_details

# jsonファイルに存在するかどうか確認する
# masterないに存在する日付データを取得する
def check_key_in_master(json_file, key, checkdatas="master"):
    with open(json_file, "r", encoding="utf-8") as file:
        json_data = json.load(file)
        master_data = json_data.get(checkdatas, {})
        for item_key, item_value in master_data.items():
            if item_key == key:
                return item_value
        # return key in master_data




# 現在の日付を取得
today_date = datetime.now().date()

# JSONファイルを読み込む
with open(BuchererMainDatasjson, "r", encoding="utf-8") as file:
    data = json.load(file)

# JSONデータ全体を走査し、"2024-02-25"形式の日付のみを抽出
valid_dates = []
for section in data.values():
    if isinstance(section, dict):
        for key in section.keys():
            try:
                datetime.strptime(key, "%Y-%m-%d")
                valid_dates.append(key)
            except ValueError:
                pass

# 実行日に最も近い日付を見つける
nearest_date = min(
    valid_dates, key=lambda x: abs(today_date - datetime.strptime(x, "%Y-%m-%d").date())
)

print("本日の日付:", today_date)
print("最も近い日付:", nearest_date)


# エクセルファイル作成
file_name = f"TOURNEAU CPO リスト{today_date}.xlsx"
if not os.path.exists(file_name):
    # Excelブックの作成
    wb = Workbook()
    ws = wb.active
    # ヘッダー行を追加
    ws.append(
        [
            "商品番号",
            "モデル",
            "年式",
            "サイズ",
            "素材",
            "Ref.",
            "ブレスレット",
            "ダイアル",
            "再販価格",
        ]
    )
else:
    # ファイルが存在する場合は既存のファイルを読み込み
    wb = load_workbook(file_name)
    ws = wb.active

with open(BuchererMainDatasjson, "r", encoding="utf-8") as file:
    json_data = json.load(file)
    nearest_date_data = json_data.get(nearest_date, {})
    print("ループチェック")

    # 日付内のアイテムをループする
    for key, value in nearest_date_data.items():
        row_items = []
        bracelet = ""
        # 日付>アイテムナンバー
        datedata = check_key_in_master(
            BuchererMainDatasjson, key=key, checkdatas=nearest_date
        )
        # 金額は、アイテムナンバーごとに保存されているので、それを取得する。
        price = datedata.get("price")

        # price = check_key_in_master(BuchererMainDatasjson, key="price", checkdatas=key)

        # json形式で返ってくる
        hitvalue = check_key_in_master(BuchererMainDatasjson, key=key)
        model = hitvalue.get("model")
        # ブレスレットが存在するか確認する。空が返ってきた場合、データを検索、それでも一致しない場合は手でいれる、、
        bracelet = hitvalue.get("bracelet")
      

        modelreplasename = check_key_in_master(
            BuchererMainDatasjson, key=model, checkdatas="modelreplasedata"
        )
        # 配列にする_i
        # モデル
        itemno = key
        # サイズ
        size = hitvalue.get("size")

        # リファレンスナンバー
        ref = hitvalue.get("ref")
        # 検索してアイテムの詳細を取得
        matching_items = get_item_details(main_json_data, ref)
        # 一致したアイテムの詳細を出力
        # for item in matching_items:
        #     # items_with_non_empty_ref = [i for i in item if item.get("ref") != ""]
        #     # item["ref"]
        #     # print(f"{item[\"ref\"]}リファレンス {item[\"bracelet\"]}")

        #     print(f"{item["ref"]}リファレンス {item["bracelet"]}")
            # print(f"{item}:はヒットした値。{ref}はリファレンスナンバー")

        # アイテムコードを返す。
        itemcodereplace = extract_single_digit(ref)
        itemcodereplace_alpahbet = check_key_in_master(
            BuchererMainDatasjson, key=itemcodereplace, checkdatas="item-code"
        )
        year = hitvalue.get("year")
        bracelet = hitvalue.get("bracelet")
        dial = hitvalue.get("dial")

        url = hitvalue.get("url")
        row_items.append(key)
        row_items.append(modelreplasename)
        row_items.append(year)
        row_items.append(size)
        row_items.append(itemcodereplace_alpahbet)
        row_items.append(ref)
        row_items.append(bracelet)
        row_items.append(dial)
        row_items.append(price)
        row_items.append(url)

        print(size, "はサイズ")
        print(year, "は年度")
        print(url, "はURL")
        print(price, "は金額")
        print(ref, "リファレンスナンバー")
        print(itemcodereplace_alpahbet, "はアイテムコード")
        print(model)
        print(modelreplasename)
        print("---------------------------------------")
        ws.append(row_items)
wb.save(file_name)
# return key in master_data
